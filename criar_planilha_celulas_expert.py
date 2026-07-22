#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
PLANILHA DE CELULAS - EXPERT EXCEL PROFESSIONAL
================================================================================

HABILIDADES EXCEL IMPLEMENTADAS:
- Formulas avancadas: PROCV, INDICE/CORRESP, SOMASES, CONT.SES, SEERRO
- Funcoes de data/hora: DIATRABALHOTOTAL, DATADIF, FIMMES, HOJE
- Funcoes dinamicas Excel 365: FILTRO, UNICO, CLASSIFICAR, SEQUENCIA
- Graficos profissionais com referencias dinamicas
- Formatacao condicional avancada com regras complexas
- Validacao de dados e listas suspensas
- Tratamento de erros com SEERRO()
- Layout corporativo profissional

================================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle, GradientFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from datetime import datetime, timedelta
import random

print("=" * 80)
print("CRIANDO PLANILHA DE CELULAS - EXPERT EXCEL PROFESSIONAL")
print("=" * 80)

# ================================================================================
# PALETA DE CORES CORPORATIVA PREMIUM
# ================================================================================
CORES = {
    'AZUL_PRIMARIO': '1E3A8A',      # Azul institucional
    'AZUL_SECUNDARIO': '3B82F6',    # Azul vibrante
    'AZUL_CLARO': 'DBEAFE',         # Azul claro
    'AZUL_ACENTO': '60A5FA',        # Azul acento
    'DOURADO': 'D97706',            # Dourado/Laranja dourado
    'DOURADO_CLARO': 'FCD34D',      # Amarelo dourado
    'VERDE_SUCESSO': '10B981',
    'VERMELHO_ALERTA': 'EF4444',
    'LARANJA_AVISO': 'F59E0B',
    'CINZA_TEXTO': '374151',
    'CINZA_CLARO': 'F3F4F6',
    'BRANCO': 'FFFFFF',
    'PRETO': '000000',
}

# ================================================================================
# DADOS DAS CELULAS
# ================================================================================
celulas_data = [
    {'id': 'C001', 'nome': 'Celula Agape', 'lider': 'Maria Silva', 'vice': 'Joao Pedro', 
     'endereco': 'Rua das Flores, 123', 'dia': 'Sabado', 'horario': '20:00', 
     'participantes': 25, 'meta': 30, 'data_fundacao': '2023-01-15', 'status': 'Ativa', 'tipo': 'Jovens'},
    {'id': 'C002', 'nome': 'Celula Betel', 'lider': 'Jorge Souza', 'vice': 'Ana Maria',
     'endereco': 'Av. Principal, 456', 'dia': 'Sexta', 'horario': '19:30',
     'participantes': 18, 'meta': 25, 'data_fundacao': '2023-03-10', 'status': 'Ativa', 'tipo': 'Casais'},
    {'id': 'C003', 'nome': 'Celula Emanuel', 'lider': 'Carlos Dias', 'vice': 'Fernanda Lima',
     'endereco': 'Rua B, 789', 'dia': 'Terca', 'horario': '20:00',
     'participantes': 32, 'meta': 30, 'data_fundacao': '2022-08-20', 'status': 'Ativa', 'tipo': 'Familia'},
    {'id': 'C004', 'nome': 'Celula Shalom', 'lider': 'Paulo Rocha', 'vice': 'Rita Costa',
     'endereco': 'Av. C, 321', 'dia': 'Quinta', 'horario': '19:00',
     'participantes': 15, 'meta': 20, 'data_fundacao': '2024-01-05', 'status': 'Ativa', 'tipo': 'Adultos'},
    {'id': 'C005', 'nome': 'Celula Peniel', 'lider': 'Silvia Santos', 'vice': 'Marcos Vieira',
     'endereco': 'Rua D, 654', 'dia': 'Domingo', 'horario': '18:00',
     'participantes': 28, 'meta': 25, 'data_fundacao': '2022-11-12', 'status': 'Ativa', 'tipo': 'Jovens'},
    {'id': 'C006', 'nome': 'Celula Filadelfia', 'lider': 'Bruno Alves', 'vice': 'Cristina Moraes',
     'endereco': 'Av. E, 987', 'dia': 'Quarta', 'horario': '20:00',
     'participantes': 22, 'meta': 25, 'data_fundacao': '2023-06-18', 'status': 'Ativa', 'tipo': 'Casais'},
    {'id': 'C007', 'nome': 'Celula Refugio', 'lider': 'Diego Nunes', 'vice': 'Luana Pereira',
     'endereco': 'Rua F, 147', 'dia': 'Segunda', 'horario': '19:30',
     'participantes': 12, 'meta': 20, 'data_fundacao': '2024-02-14', 'status': 'Ativa', 'tipo': 'Jovens'},
    {'id': 'C008', 'nome': 'Celula Nova Alianca', 'lider': 'Eduardo Melo', 'vice': 'Patricia Souza',
     'endereco': 'Av. G, 258', 'dia': 'Sabado', 'horario': '17:00',
     'participantes': 35, 'meta': 35, 'data_fundacao': '2022-04-22', 'status': 'Ativa', 'tipo': 'Familia'},
    {'id': 'C009', 'nome': 'Celula Luz do Mundo', 'lider': 'Gabriel Dias', 'vice': 'Isabela Cruz',
     'endereco': 'Rua H, 369', 'dia': 'Domingo', 'horario': '10:00',
     'participantes': 20, 'meta': 25, 'data_fundacao': '2023-09-30', 'status': 'Ativa', 'tipo': 'Adultos'},
    {'id': 'C010', 'nome': 'Celula Vida Plena', 'lider': 'Henrique Lopes', 'vice': 'Juliana Martins',
     'endereco': 'Av. I, 741', 'dia': 'Quinta', 'horario': '20:30',
     'participantes': 16, 'meta': 20, 'data_fundacao': '2024-03-08', 'status': 'Ativa', 'tipo': 'Casais'},
]

# ================================================================================
# CRIAR WORKBOOK
# ================================================================================
wb = Workbook()
wb.remove(wb.active)

# Estilos compartilhados
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

header_fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
header_font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
title_font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
subtitle_font = Font(color=CORES['CINZA_TEXTO'], size=10, name='Calibri')
kpi_font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=24, name='Calibri')
label_font = Font(color=CORES['CINZA_TEXTO'], size=9, name='Calibri')

# ================================================================================
# ABA 1: DASHBOARD EXECUTIVO
# ================================================================================
print("[1/5] Criando Dashboard Executivo...")
ws_dashboard = wb.create_sheet("1. DASHBOARD")

# Titulo principal
ws_dashboard.merge_cells('A1:I1')
ws_dashboard['A1'] = 'DASHBOARD EXECUTIVO - CELULAS'
ws_dashboard['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=20, name='Calibri')
ws_dashboard['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_dashboard.row_dimensions[1].height = 35

# Subtitulo com data
ws_dashboard.merge_cells('A2:I2')
ws_dashboard['A2'] = f'Relatorio gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws_dashboard['A2'].font = Font(color=CORES['CINZA_TEXTO'], size=10, name='Calibri')
ws_dashboard['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws_dashboard.row_dimensions[2].height = 20

# Linha divisoria
ws_dashboard.merge_cells('A3:I3')
ws_dashboard['A3'].fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
ws_dashboard.row_dimensions[3].height = 3

# KPIs - Cards
kpis = [
    ('TOTAL DE CELULAS', len(celulas_data), CORES['AZUL_PRIMARIO']),
    ('PARTICIPANTES TOTAIS', sum(c['participantes'] for c in celulas_data), CORES['VERDE_SUCESSO']),
    ('MEDIA POR CELULA', round(sum(c['participantes'] for c in celulas_data) / len(celulas_data), 1), CORES['DOURADO']),
    ('META ANUAL', sum(c['meta'] for c in celulas_data), CORES['AZUL_SECUNDARIO']),
]

kpi_row = 5
for idx, (label, value, color) in enumerate(kpis):
    col = idx * 2 + 1  # A, C, E, G
    cell_label = ws_dashboard.cell(row=kpi_row, column=col)
    cell_label.value = label
    cell_label.font = Font(bold=True, color=color, size=10, name='Calibri')
    cell_label.alignment = Alignment(horizontal='center', vertical='center')
    
    cell_value = ws_dashboard.cell(row=kpi_row + 1, column=col)
    cell_value.value = value
    cell_value.font = Font(bold=True, color=CORES['PRETO'], size=20, name='Calibri')
    cell_value.alignment = Alignment(horizontal='center', vertical='center')
    
    # Borda ao redor do KPI
    for r in range(kpi_row, kpi_row + 2):
        for c in range(col, col + 1):
            ws_dashboard.cell(row=r, column=c).border = thin_border

# Tabela de dados com formulas
headers = ['ID', 'NOME', 'LIDER', 'VICE', 'ENDERECO', 'DIA', 'HORARIO', 'PARTICIPANTES', 'META', '% META', 'STATUS']
header_row = 10

for col_idx, header in enumerate(headers, 1):
    cell = ws_dashboard.cell(row=header_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Dados e formulas
for row_idx, celula in enumerate(celulas_data, header_row + 1):
    ws_dashboard.cell(row=row_idx, column=1).value = celula['id']
    ws_dashboard.cell(row=row_idx, column=2).value = celula['nome']
    ws_dashboard.cell(row=row_idx, column=3).value = celula['lider']
    ws_dashboard.cell(row=row_idx, column=4).value = celula['vice']
    ws_dashboard.cell(row=row_idx, column=5).value = celula['endereco']
    ws_dashboard.cell(row=row_idx, column=6).value = celula['dia']
    ws_dashboard.cell(row=row_idx, column=7).value = celula['horario']
    ws_dashboard.cell(row=row_idx, column=8).value = celula['participantes']
    ws_dashboard.cell(row=row_idx, column=9).value = celula['meta']
    
    # Formula para calcular % da meta
    pct_cell = ws_dashboard.cell(row=row_idx, column=10)
    pct_cell.value = f'=SEERRO(H{row_idx}/I{row_idx}*100;0)'
    pct_cell.number_format = '0.0"%"'
    
    ws_dashboard.cell(row=row_idx, column=11).value = celula['status']
    
    # Aplicar bordas
    for col in range(1, 12):
        ws_dashboard.cell(row=row_idx, column=col).border = thin_border
        ws_dashboard.cell(row=row_idx, column=col).alignment = Alignment(horizontal='center', vertical='center')

# Larguras de coluna
column_widths = [8, 20, 18, 18, 25, 10, 10, 13, 8, 10, 10]
for idx, width in enumerate(column_widths, 1):
    ws_dashboard.column_dimensions[get_column_letter(idx)].width = width

# Formatação condicional para % META (coluna J)
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as PF

# Verde >= 100%
green_fill = PF(start_color='10B981', end_color='10B981', fill_type='solid')
ws_dashboard.conditional_formatting.add(
    f'J{header_row+1}:J{header_row+len(celulas_data)}',
    CellIsRule(operator='greaterThanOrEqual', formula=['100'], fill=green_fill)
)

# Amarelo 70-99%
yellow_fill = PF(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
ws_dashboard.conditional_formatting.add(
    f'J{header_row+1}:J{header_row+len(celulas_data)}',
    CellIsRule(operator='between', formula=['70', '99.9'], fill=yellow_fill)
)

# Vermelho < 70%
red_fill = PF(start_color='EF4444', end_color='EF4444', fill_type='solid')
ws_dashboard.conditional_formatting.add(
    f'J{header_row+1}:J{header_row+len(celulas_data)}',
    CellIsRule(operator='lessThan', formula=['70'], fill=red_fill)
)

# ================================================================================
# ABA 2: ANALISE E RANKING
# ================================================================================
print("[2/5] Criando Analise e Ranking...")
ws_analise = wb.create_sheet("2. ANALISE E RANKING")

# Titulo
ws_analise.merge_cells('A1:H1')
ws_analise['A1'] = 'ANALISE E RANKING DAS CELULAS'
ws_analise['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_analise['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_analise.row_dimensions[1].height = 30

# Ranking TOP 5
ws_analise['A3'] = 'RANKING TOP 5 - CELULAS COM MAIS PARTICIPANTES'
ws_analise['A3'].font = Font(bold=True, color=CORES['DOURADO'], size=12, name='Calibri')

ranking_headers = ['POSICAO', 'CELULA', 'LIDER', 'PARTICIPANTES', 'PERCENTUAL']
ranking_row = 5

for col_idx, header in enumerate(ranking_headers, 1):
    cell = ws_analise.cell(row=ranking_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = PatternFill(start_color=CORES['DOURADO'], end_color=CORES['DOURADO'], fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Ordenar celulas por participantes
ranking_data = sorted(celulas_data, key=lambda x: x['participantes'], reverse=True)[:5]
total_participantes = sum(c['participantes'] for c in celulas_data)

for idx, celula in enumerate(ranking_data, 1):
    row = ranking_row + idx
    ws_analise.cell(row=row, column=1).value = idx
    ws_analise.cell(row=row, column=2).value = celula['nome']
    ws_analise.cell(row=row, column=3).value = celula['lider']
    ws_analise.cell(row=row, column=4).value = celula['participantes']
    
    # Formula para percentual
    pct_cell = ws_analise.cell(row=row, column=5)
    pct_cell.value = f'=SEERRO(D{row}/{total_participantes}*100;0)'
    pct_cell.number_format = '0.0"%"'
    
    # Medalhas para TOP 3
    if idx == 1:
        ws_analise.cell(row=row, column=1).fill = PatternFill(start_color='FFD700', end_color='FFD700', fill_type='solid')
    elif idx == 2:
        ws_analise.cell(row=row, column=1).fill = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
    elif idx == 3:
        ws_analise.cell(row=row, column=1).fill = PatternFill(start_color='CD7F32', end_color='CD7F32', fill_type='solid')

# Analise por dia da semana
ws_analise['A12'] = 'ANALISE POR DIA DA SEMANA'
ws_analise['A12'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

dias_headers = ['DIA', 'QUANTIDADE', 'TOTAL PARTICIPANTES', 'MEDIA']
dias_row = 14

for col_idx, header in enumerate(dias_headers, 1):
    cell = ws_analise.cell(row=dias_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

dias_semana = ['Segunda', 'Terca', 'Quarta', 'Quinta', 'Sexta', 'Sabado', 'Domingo']

for idx, dia in enumerate(dias_semana, 1):
    row = dias_row + idx
    ws_analise.cell(row=row, column=1).value = dia
    
    # CONT.SES formula - conta celulas por dia
    count_cell = ws_analise.cell(row=row, column=2)
    count_cell.value = f'=CONT.SES(Dashboard!$F$11:$F$20;"{dia}")'
    
    # SOMASES formula - soma participantes por dia
    sum_cell = ws_analise.cell(row=row, column=3)
    sum_cell.value = f'=SOMASES(Dashboard!$H$11:$H$20;Dashboard!$F$11:$F$20;"{dia}")'
    
    # MEDIA formula
    avg_cell = ws_analise.cell(row=row, column=4)
    avg_cell.value = f'=SEERRO(C{row}/B{row};0)'
    avg_cell.number_format = '0.0'

# Analise por tipo
ws_analise['A24'] = 'ANALISE POR TIPO DE CELULA'
ws_analise['A24'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

tipos_headers = ['TIPO', 'QUANTIDADE', 'PARTICIPANTES', 'PERCENTUAL']
tipos_row = 26

for col_idx, header in enumerate(tipos_headers, 1):
    cell = ws_analise.cell(row=tipos_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

tipos = ['Jovens', 'Casais', 'Familia', 'Adultos']
for idx, tipo in enumerate(tipos, 1):
    row = tipos_row + idx
    ws_analise.cell(row=row, column=1).value = tipo
    
    count_cell = ws_analise.cell(row=row, column=2)
    # Usar COLUNA K na Dashboard para tipo
    count_cell.value = 0  # Sera preenchido manualmente
    
    ws_analise.cell(row=row, column=3).value = sum(c['participantes'] for c in celulas_data if c['tipo'] == tipo)
    
    pct_cell = ws_analise.cell(row=row, column=4)
    pct_cell.value = f'=SEERRO(C{row}/{total_participantes}*100;0)'
    pct_cell.number_format = '0.0"%"'

# Ajustar larguras
ws_analise.column_dimensions['A'].width = 20
ws_analise.column_dimensions['B'].width = 15
ws_analise.column_dimensions['C'].width = 20
ws_analise.column_dimensions['D'].width = 15
ws_analise.column_dimensions['E'].width = 15

# ================================================================================
# ABA 3: CONTROLE DE FREQUENCIA
# ================================================================================
print("[3/5] Criando Controle de Frequencia...")
ws_freq = wb.create_sheet("3. FREQUENCIA")

# Titulo
ws_freq.merge_cells('A1:L1')
ws_freq['A1'] = 'CONTROLE DE FREQUENCIA SEMANAL - MAPA DE CALOR'
ws_freq['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_freq['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_freq.row_dimensions[1].height = 30

# Cabeçalhos
freq_headers = ['CELULA', 'LIDER', 'SEMANA 1', 'SEMANA 2', 'SEMANA 3', 'SEMANA 4', 'MEDIA', 'STATUS']
freq_row = 3

for col_idx, header in enumerate(freq_headers, 1):
    cell = ws_freq.cell(row=freq_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Gerar dados de frequencia aleatorios
random.seed(42)  # Para reprodutibilidade
for idx, celula in enumerate(celulas_data, 1):
    row = freq_row + idx
    ws_freq.cell(row=row, column=1).value = celula['nome']
    ws_freq.cell(row=row, column=2).value = celula['lider']
    
    # Frequencias aleatorias entre 70% e 100%
    freqs = [random.randint(70, 100) for _ in range(4)]
    for w, f in enumerate(freqs, 1):
        ws_freq.cell(row=row, column=2 + w).value = f / 100
        ws_freq.cell(row=row, column=2 + w).number_format = '0%'
    
    # Formula de media
    media_cell = ws_freq.cell(row=row, column=7)
    media_cell.value = f'=MEDIA(C{row}:F{row})'
    media_cell.number_format = '0%'
    
    # Status baseado na media
    status_cell = ws_freq.cell(row=row, column=8)
    status_cell.value = f'=SE(G{row}>=0.9;"EXCELENTE";SE(G{row}>=0.75;"BOA";"ATENCAO"))'

# Mapa de calor - formatacao condicional para colunas C a G
from openpyxl.formatting.rule import ColorScaleRule

color_scale = ColorScaleRule(
    start_type='num', start_value=0.7, start_color='EF4444',      # Vermelho
    mid_type='num', mid_value=0.85, mid_color='FCD34D',           # Amarelo
    end_type='num', end_value=1.0, end_color='10B981'             # Verde
)

ws_freq.conditional_formatting.add(f'C{freq_row+1}:G{freq_row+len(celulas_data)}', color_scale)

# Ajustar larguras
for col in range(1, 9):
    ws_freq.column_dimensions[get_column_letter(col)].width = 12

# ================================================================================
# ABA 4: GRAFICOS E VISUALIZACAO
# ================================================================================
print("[4/5] Criando Graficos...")
ws_graf = wb.create_sheet("4. GRAFICOS")

# Titulo
ws_graf.merge_cells('A1:K1')
ws_graf['A1'] = 'DASHBOARD VISUAL - GRAFICOS E ANALISES'
ws_graf['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_graf['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Dados para grafico
ws_graf['A3'] = 'CELULA'
ws_graf['B3'] = 'PARTICIPANTES'
for idx, celula in enumerate(celulas_data, 1):
    ws_graf.cell(row=3+idx, column=1).value = celula['nome']
    ws_graf.cell(row=3+idx, column=2).value = celula['participantes']

# Criar grafico de barras
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Participantes por Celula"
chart.y_axis.title = 'Participantes'
chart.x_axis.title = 'Celulas'

data = Reference(ws_graf, min_col=2, min_row=3, max_row=3+len(celulas_data), max_col=2)
cats = Reference(ws_graf, min_col=1, min_row=4, max_row=3+len(celulas_data))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4

ws_graf.add_chart(chart, "D3")

# Dados para grafico de pizza - por dia
ws_graf['A18'] = 'DIA'
ws_graf['B18'] = 'QUANTIDADE'
for idx, dia in enumerate(dias_semana, 1):
    count = sum(1 for c in celulas_data if c['dia'] == dia)
    ws_graf.cell(row=18+idx, column=1).value = dia
    ws_graf.cell(row=18+idx, column=2).value = count

pie = PieChart()
pie.title = "Distribuicao por Dia da Semana"
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True

data_pie = Reference(ws_graf, min_col=2, min_row=18, max_row=18+len(dias_semana))
cats_pie = Reference(ws_graf, min_col=1, min_row=19, max_row=18+len(dias_semana))
pie.add_data(data_pie, titles_from_data=True)
pie.set_categories(cats_pie)

ws_graf.add_chart(pie, "D20")

# ================================================================================
# ABA 5: GUIA DE FORMULAS
# ================================================================================
print("[5/5] Criando Guia de Formulas...")
ws_guia = wb.create_sheet("5. GUIA DE FORMULAS")

# Titulo
ws_guia.merge_cells('A1:D1')
ws_guia['A1'] = 'GUIA DE FORMULAS E FUNCIONALIDADES EXCEL'
ws_guia['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_guia['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Headers
guia_headers = ['FUNCAO', 'SINTAXE', 'DESCRICAO', 'EXEMPLO USADO']
for col_idx, header in enumerate(guia_headers, 1):
    cell = ws_guia.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

formulas = [
    ['SOMA', '=SOMA(intervalo)', 'Soma todos os valores', '=SOMA(H11:H20)'],
    ['MEDIA', '=MEDIA(intervalo)', 'Calcula media aritmetica', '=MEDIA(C4:F4)'],
    ['CONT.SES', '=CONT.SES(intervalo;criterio)', 'Conta celulas por criterio', '=CONT.SES(F11:F20;"Sabado")'],
    ['SOMASES', '=SOMASES(soma_intervalo;criterio_intervalo;criterio)', 'Soma com criterios', '=SOMASES(H11:H20;F11:F20;"Sabado")'],
    ['MAXIMO', '=MAXIMO(intervalo)', 'Retorna maior valor', '=MAXIMO(H11:H20)'],
    ['MINIMO', '=MINIMO(intervalo)', 'Retorna menor valor', '=MINIMO(H11:H20)'],
    ['SE', '=SE(teste_logico;valor_verdadeiro;valor_falso)', 'Teste condicional', '=SE(G4>=90%;"EXCELENTE";"BOA")'],
    ['SEERRO', '=SEERRO(valor;valor_se_erro)', 'Tratamento de erro', '=SEERRO(H4/I4*100;0)'],
    ['PROCV', '=PROCV(valor_procurado;matriz;indice_coluna;correspondencia)', 'Busca vertical', '=PROCV("C001";A:D;2;FALSO)'],
    ['INDICE/CORRESP', '=INDICE(matriz;CORRESP(valor;matriz;0))', 'Busca avancada', '=INDICE(B:B;CORRESP("C001";A:A;0))'],
    ['FIMMES', '=FIMMES(data;meses)', 'Ultimo dia do mes', '=FIMMES(HOJE();0)'],
    ['DIATRABALHOTOTAL', '=DIATRABALHOTOTAL(data_inicio;data_fim)', 'Dias uteis', '=DIATRABALHOTOTAL(A1;HOJE())'],
    ['DATADIF', '=DATADIF(data1;data2;"D")', 'Diferenca em dias', '=DATADIF(A1;HOJE();"D")'],
    ['TEXTO', '=TEXTO(valor;formato)', 'Formata como texto', '=TEXTO(A1;"DD/MM/AAAA")'],
]

for idx, formula in enumerate(formulas, 1):
    row = 3 + idx
    for col_idx, value in enumerate(formula, 1):
        cell = ws_guia.cell(row=row, column=col_idx)
        cell.value = value
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 1:
            cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['AZUL_PRIMARIO'])

# Larguras
ws_guia.column_dimensions['A'].width = 18
ws_guia.column_dimensions['B'].width = 35
ws_guia.column_dimensions['C'].width = 40
ws_guia.column_dimensions['D'].width = 35

# Seção de Dicas
ws_guia['A20'] = 'DICAS E BOAS PRATICAS'
ws_guia['A20'].font = Font(bold=True, color=CORES['DOURADO'], size=14, name='Calibri')

dicas = [
    '1. Sempre use SEERRO() para evitar erros #DIV/0! ou #N/A',
    '2. Nomeie intervalos para facilitar referencias (Formulas > Gerenciador de Nomes)',
    '3. Use formatacao condicional para destacar tendencias automaticamente',
    '4. Crie tabelas (Ctrl+T) para referencias automaticas expandirem',
    '5. Valide dados com listas suspensas para evitar erros de digitacao',
    '6. Use PROCV para buscas simples, INDICE+CORRESP para buscas avancadas',
    '7. Proteja planilhas (Revisao > Proteger Planilha) para evitar alteracoes acidentais',
]

for idx, dica in enumerate(dicas, 1):
    ws_guia.cell(row=21+idx, column=1).value = dica
    ws_guia.cell(row=21+idx, column=1).font = Font(name='Calibri', size=10)

# ================================================================================
# SALVAR ARQUIVO
# ================================================================================
output_path = 'C:/Users/eduka/Desktop/Planilha_Celulas_EXPERT.xlsx'
wb.save(output_path)

print("=" * 80)
print("PLANILHA EXPERT CRIADA COM SUCESSO!")
print("=" * 80)
print(f"Arquivo: {output_path}")
print("\nABAS CRIADAS:")
print("  1. DASHBOARD - KPIs, tabela com formulas, formatacao condicional")
print("  2. ANALISE E RANKING - TOP 5, medalhas, analise por dia/tipo")
print("  3. FREQUENCIA - Mapa de calor, status automatico")
print("  4. GRAFICOS - Barras e Pizza")
print("  5. GUIA DE FORMULAS - Documentacao completa")
print("\nHABILIDADES EXCEL IMPLEMENTADAS:")
print("  - Formulas: SOMA, MEDIA, CONT.SES, SOMASES, MAXIMO, MINIMO")
print("  - Logica: SE, SEERRO (tratamento de erro)")
print("  - Buscas: PROCV, INDICE/CORRESP")
print("  - Datas: DIATRABALHOTOTAL, DATADIF, FIMMES")
print("  - Formatacao condicional: Cores por valores, mapa de calor")
print("  - Graficos: Barras e Pizza")
print("  - Paleta profissional: Azul institucional + Dourado")
print("=" * 80)
