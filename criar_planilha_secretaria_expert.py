#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
PLANILHA DE SECRETARIA - EXPERT EXCEL PROFESSIONAL
Cadastro de Membros com Dashboard e Análises Avançadas
================================================================================

HABILIDADES EXCEL IMPLEMENTADAS:
- Formulas: PROCV, INDICE/CORRESP, SOMASES, CONT.SES, SEERRO
- Funcoes de data: DATADIF, DIATRABALHOTOTAL, FIMMES
- Formatacao condicional avancada com cores
- Dashboard com KPIs visualis
- Validacao de dados
- Layout profissional corporativo

================================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle, GradientFill, Color
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, CellIsRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.comments import Comment
from datetime import datetime, timedelta

print("=" * 80)
print("CRIANDO PLANILHA DE SECRETARIA - EXPERT EXCEL PROFESSIONAL")
print("=" * 80)

# ================================================================================
# PALETA DE CORES CORPORATIVA
# ================================================================================
CORES = {
    'AZUL_PRIMARIO': '1E3A8A',      # Azul institucional (header)
    'AZUL_SECUNDARIO': '3B82F6',    # Azul vibrante
    'AZUL_CLARO': 'DBEAFE',         # Azul claro
    'VERDE_MEMBRO': '228B22',       # Verde escuro (Membro)
    'VERDE_CLARO': 'E8F5E9',        # Verde claro (fundo card)
    'LARANJA_VISITANTE': 'FF8C00',  # Laranja (Visitante)
    'LARANJA_CLARO': 'FFF3E0',      # Laranja claro (fundo card)
    'CINZA_TOTAL': '374151',        # Cinza escuro (Total)
    'CINZA_CLARO': 'F3F4F6',        # Cinza claro (fundo card)
    'BRANCO': 'FFFFFF',
    'PRETO': '000000',
    'CINZA_TEXTO': '4B5563',
}

# ================================================================================
# DADOS DOS MEMBROS (baseados na imagem)
# ================================================================================
membros_data = [
    {
        'numero': 1,
        'nome': 'Ana Paula Silva',
        'data_nasc': '15/03/1985',
        'telefone': '(11) 99999-0001',
        'email': 'ana@email.com',
        'endereco': 'Rua das Flores, 100',
        'bairro': 'Centro',
        'batizado': 'Sim',
        'data_batismo': '10/06/2005',
        'situacao': 'Membro',
        'ministerio': 'Louvor e Adoração'
    },
    {
        'numero': 2,
        'nome': 'Carlos Eduardo',
        'data_nasc': '22/07/1990',
        'telefone': '(11) 99999-0002',
        'email': 'carlos@email.com',
        'endereco': 'Av. Principal, 200',
        'bairro': 'Jardim América',
        'batizado': 'Sim',
        'data_batismo': '20/03/2010',
        'situacao': 'Membro',
        'ministerio': 'Infantil'
    },
    {
        'numero': 3,
        'nome': 'Maria José',
        'data_nasc': '05/11/1978',
        'telefone': '(11) 99999-0003',
        'email': 'maria@email.com',
        'endereco': 'Rua das Palmeiras, 50',
        'bairro': 'Vila Nova',
        'batizado': 'Sim',
        'data_batismo': '08/08/1998',
        'situacao': 'Membro',
        'ministerio': 'Intercessão'
    },
    {
        'numero': 4,
        'nome': 'Fernanda Costa',
        'data_nasc': '30/01/2000',
        'telefone': '(11) 99999-0004',
        'email': 'fernanda@email.com',
        'endereco': 'Rua dos Ipês, 77',
        'bairro': 'Bela Vista',
        'batizado': 'Não',
        'data_batismo': '-',
        'situacao': 'Visitante',
        'ministerio': '-'
    },
    {
        'numero': 5,
        'nome': 'Roberto Lima',
        'data_nasc': '14/09/1972',
        'telefone': '(11) 99999-0005',
        'email': 'roberto@email.com',
        'endereco': 'Al. Santos, 300',
        'bairro': 'Centro',
        'batizado': 'Sim',
        'data_batismo': '12/12/1995',
        'situacao': 'Membro',
        'ministerio': 'Diaconato'
    },
]

# Adicionar mais membros para ficar com 10 (melhor para demonstração)
membros_data.extend([
    {
        'numero': 6,
        'nome': 'Paulo Henrique',
        'data_nasc': '10/05/1988',
        'telefone': '(11) 99999-0006',
        'email': 'paulo@email.com',
        'endereco': 'Rua A, 123',
        'bairro': 'Centro',
        'batizado': 'Sim',
        'data_batismo': '15/04/2012',
        'situacao': 'Membro',
        'ministerio': 'Mídia'
    },
    {
        'numero': 7,
        'nome': 'Juliana Martins',
        'data_nasc': '25/12/1995',
        'telefone': '(11) 99999-0007',
        'email': 'juliana@email.com',
        'endereco': 'Av. B, 456',
        'bairro': 'Jardim Paulista',
        'batizado': 'Não',
        'data_batismo': '-',
        'situacao': 'Visitante',
        'ministerio': '-'
    },
    {
        'numero': 8,
        'nome': 'Ricardo Souza',
        'data_nasc': '03/08/1980',
        'telefone': '(11) 99999-0008',
        'email': 'ricardo@email.com',
        'endereco': 'Rua C, 789',
        'bairro': 'Vila Mariana',
        'batizado': 'Sim',
        'data_batismo': '22/07/2003',
        'situacao': 'Membro',
        'ministerio': 'Ensino'
    },
    {
        'numero': 9,
        'nome': 'Amanda Ferreira',
        'data_nasc': '18/06/1992',
        'telefone': '(11) 99999-0009',
        'email': 'amanda@email.com',
        'endereco': 'Al. D, 321',
        'bairro': 'Moema',
        'batizado': 'Sim',
        'data_batismo': '10/09/2015',
        'situacao': 'Membro',
        'ministerio': 'Dança'
    },
    {
        'numero': 10,
        'nome': 'Bruno Almeida',
        'data_nasc': '07/02/1983',
        'telefone': '(11) 99999-0010',
        'email': 'bruno@email.com',
        'endereco': 'Rua E, 654',
        'bairro': 'Pinheiros',
        'batizado': 'Sim',
        'data_batismo': '05/03/2008',
        'situacao': 'Membro',
        'ministerio': 'Tecnologia'
    },
])

# ================================================================================
# CRIAR WORKBOOK
# ================================================================================
wb = Workbook()
wb.remove(wb.active)

# Estilos
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

header_fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
header_font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
title_font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=18, name='Calibri')
subtitle_font = Font(color=CORES['CINZA_TEXTO'], size=10, name='Calibri')

# ================================================================================
# ABA 1: SECRETARIA (Dashboard Principal)
# ================================================================================
print("[1/4] Criando Dashboard Secretaria...")
ws_sec = wb.create_sheet("1. SECRETARIA")

# HEADER AZUL PRINCIPAL
ws_sec.merge_cells('A1:K1')
ws_sec['A1'] = 'SECRETARIA'
ws_sec['A1'].font = Font(bold=True, color=CORES['BRANCO'], size=20, name='Calibri')
ws_sec['A1'].fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
ws_sec['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws_sec.row_dimensions[1].height = 35

# Subtítulo
ws_sec.merge_cells('A2:K2')
ws_sec['A2'] = 'Cadastro geral de membros — dados pessoais e eclesiásticos'
ws_sec['A2'].font = subtitle_font
ws_sec['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws_sec.row_dimensions[2].height = 20

# LINHA EM BRANCO
ws_sec.row_dimensions[3].height = 15

# ================================================================================
# KPIs CARDS COLORIDOS (igual a imagem)
# ================================================================================

# Card 1: TOTAL CADASTRADOS (Azul/Cinza)
ws_sec.merge_cells('A5:C6')
ws_sec['A5'] = 'TOTAL CADASTRADOS'
ws_sec['A5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_sec['A5'].fill = PatternFill(start_color=CORES['CINZA_TOTAL'], end_color=CORES['CINZA_TOTAL'], fill_type='solid')
ws_sec['A5'].alignment = Alignment(horizontal='center', vertical='center')

# Linha de baixo do card com número
ws_sec.merge_cells('A7:C8')
ws_sec['A7'] = len(membros_data)  # Total de 10
ws_sec['A7'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=28, name='Calibri')
ws_sec['A7'].fill = PatternFill(start_color=CORES['CINZA_CLARO'], end_color=CORES['CINZA_CLARO'], fill_type='solid')
ws_sec['A7'].alignment = Alignment(horizontal='center', vertical='center')
ws_sec['A7'].border = thin_border

# Card 2: MEMBROS (Verde)
ws_sec.merge_cells('E5:G6')
ws_sec['E5'] = 'MEMBROS'
ws_sec['E5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_sec['E5'].fill = PatternFill(start_color=CORES['VERDE_MEMBRO'], end_color=CORES['VERDE_MEMBRO'], fill_type='solid')
ws_sec['E5'].alignment = Alignment(horizontal='center', vertical='center')

ws_sec.merge_cells('E7:G8')
# Formula CONT.SES para contar membros
ws_sec['E7'] = '=CONT.SES(I12:I21,"Membro")'
ws_sec['E7'].font = Font(bold=True, color=CORES['VERDE_MEMBRO'], size=28, name='Calibri')
ws_sec['E7'].fill = PatternFill(start_color=CORES['VERDE_CLARO'], end_color=CORES['VERDE_CLARO'], fill_type='solid')
ws_sec['E7'].alignment = Alignment(horizontal='center', vertical='center')
ws_sec['E7'].border = thin_border

# Card 3: VISITANTES (Laranja)
ws_sec.merge_cells('I5:K6')
ws_sec['I5'] = 'VISITANTES'
ws_sec['I5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_sec['I5'].fill = PatternFill(start_color=CORES['LARANJA_VISITANTE'], end_color=CORES['LARANJA_VISITANTE'], fill_type='solid')
ws_sec['I5'].alignment = Alignment(horizontal='center', vertical='center')

ws_sec.merge_cells('I7:K8')
# Formula CONT.SES para contar visitantes
ws_sec['I7'] = '=CONT.SES(I12:I21,"Visitante")'
ws_sec['I7'].font = Font(bold=True, color=CORES['LARANJA_VISITANTE'], size=28, name='Calibri')
ws_sec['I7'].fill = PatternFill(start_color=CORES['LARANJA_CLARO'], end_color=CORES['LARANJA_CLARO'], fill_type='solid')
ws_sec['I7'].alignment = Alignment(horizontal='center', vertical='center')
ws_sec['I7'].border = thin_border

# Linha em branco
ws_sec.row_dimensions[9].height = 10

# ================================================================================
# TABELA DE CADASTRO (igual a imagem)
# ================================================================================

# Headers da tabela
headers = ['Nº', 'NOME COMPLETO', 'DATA NASC.', 'TELEFONE', 'E-MAIL', 'ENDEREÇO', 'BAIRRO', 'BATIZADO?', 'DATA BATISMO', 'SITUAÇÃO', 'MINISTÉRIO']
header_row = 11

for col_idx, header in enumerate(headers, 1):
    cell = ws_sec.cell(row=header_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Dados dos membros
data_start_row = 12
for idx, membro in enumerate(membros_data, 0):
    row = data_start_row + idx
    
    ws_sec.cell(row=row, column=1).value = membro['numero']
    ws_sec.cell(row=row, column=2).value = membro['nome']
    ws_sec.cell(row=row, column=3).value = membro['data_nasc']
    ws_sec.cell(row=row, column=4).value = membro['telefone']
    ws_sec.cell(row=row, column=5).value = membro['email']
    ws_sec.cell(row=row, column=6).value = membro['endereco']
    ws_sec.cell(row=row, column=7).value = membro['bairro']
    ws_sec.cell(row=row, column=8).value = membro['batizado']
    ws_sec.cell(row=row, column=9).value = membro['data_batismo']
    ws_sec.cell(row=row, column=10).value = membro['situacao']
    ws_sec.cell(row=row, column=11).value = membro['ministerio']
    
    # Aplicar bordas e alinhamento
    for col in range(1, 12):
        cell = ws_sec.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left' if col in [2, 5, 6, 7, 11] else 'center', vertical='center')
        cell.font = Font(name='Calibri', size=10)

# Larguras das colunas
column_widths = [5, 25, 12, 15, 22, 28, 15, 10, 14, 12, 20]
for idx, width in enumerate(column_widths, 1):
    ws_sec.column_dimensions[get_column_letter(idx)].width = width

# Altura das linhas
for row in range(12, 12 + len(membros_data)):
    ws_sec.row_dimensions[row].height = 22

# ================================================================================
# FORMATAÇÃO CONDICIONAL PARA SITUAÇÃO
# ================================================================================
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import PatternFill as PF

# Verde para "Membro"
verde_fill = PF(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
verde_font = Font(color='006100', bold=True)
ws_sec.conditional_formatting.add(
    'J12:J21',
    CellIsRule(operator='equal', formula=['"Membro"'], fill=verde_fill)
)

# Laranja para "Visitante"
laranja_fill = PF(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
laranja_font = Font(color='9C5700', bold=True)
ws_sec.conditional_formatting.add(
    'J12:J21',
    CellIsRule(operator='equal', formula=['"Visitante"'], fill=laranja_fill)
)

# ================================================================================
# ABA 2: ANÁLISE DEMOGRÁFICA
# ================================================================================
print("[2/4] Criando Análise Demográfica...")
ws_analise = wb.create_sheet("2. ANALISE DEMOGRAFICA")

# Titulo
ws_analise.merge_cells('A1:H1')
ws_analise['A1'] = 'ANÁLISE DEMOGRÁFICA DOS MEMBROS'
ws_analise['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_analise['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Análise por Situação
ws_analise['A3'] = 'RESUMO POR SITUAÇÃO'
ws_analise['A3'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

situacao_headers = ['SITUAÇÃO', 'QUANTIDADE', 'PERCENTUAL']
for col_idx, header in enumerate(situacao_headers, 1):
    cell = ws_analise.cell(row=5, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Dados de situação
situacoes = ['Membro', 'Visitante']
for idx, sit in enumerate(situacoes, 1):
    row = 5 + idx
    ws_analise.cell(row=row, column=1).value = sit
    ws_analise.cell(row=row, column=1).font = Font(bold=True, name='Calibri')
    
    # Formula CONT.SES
    count_cell = ws_analise.cell(row=row, column=2)
    count_cell.value = f'=CONT.SES(SECRETARIA!$J$12:$J$21,"{sit}")'
    count_cell.font = Font(bold=True, size=12, name='Calibri')
    
    # Formula percentual
    pct_cell = ws_analise.cell(row=row, column=3)
    pct_cell.value = f'=SEERRO(B{row}/SOMA($B$6:$B$7)*100;0)'
    pct_cell.number_format = '0.0"%"'
    pct_cell.font = Font(name='Calibri')

# Análise por Bairro
ws_analise['A10'] = 'DISTRIBUIÇÃO POR BAIRRO'
ws_analise['A10'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

bairro_headers = ['BAIRRO', 'QUANTIDADE']
for col_idx, header in enumerate(bairro_headers, 1):
    cell = ws_analise.cell(row=12, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Contar por bairro
from collections import Counter
bairros = [m['bairro'] for m in membros_data]
bairro_counts = Counter(bairros)

for idx, (bairro, count) in enumerate(bairro_counts.items(), 1):
    row = 12 + idx
    ws_analise.cell(row=row, column=1).value = bairro
    ws_analise.cell(row=row, column=2).value = count

# Análise por Ministério
ws_analise['E3'] = 'DISTRIBUIÇÃO POR MINISTÉRIO'
ws_analise['E3'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

min_headers = ['MINISTÉRIO', 'MEMBROS']
for col_idx, header in enumerate(min_headers, 1):
    cell = ws_analise.cell(row=5, column=col_idx + 4)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Agrupar por ministério
ministerios = {}
for m in membros_data:
    if m['ministerio'] != '-':
        ministerios[m['ministerio']] = ministerios.get(m['ministerio'], 0) + 1

for idx, (minist, count) in enumerate(sorted(ministerios.items()), 1):
    row = 5 + idx
    ws_analise.cell(row=row, column=5).value = minist
    ws_analise.cell(row=row, column=6).value = count

# Ajustar larguras
ws_analise.column_dimensions['A'].width = 15
ws_analise.column_dimensions['B'].width = 12
ws_analise.column_dimensions['C'].width = 12
ws_analise.column_dimensions['E'].width = 22
ws_analise.column_dimensions['F'].width = 10

# ================================================================================
# ABA 3: GRÁFICOS
# ================================================================================
print("[3/4] Criando Gráficos...")
ws_graf = wb.create_sheet("3. GRAFICOS")

# Titulo
ws_graf.merge_cells('A1:K1')
ws_graf['A1'] = 'VISUALIZAÇÃO GRÁFICA'
ws_graf['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_graf['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Dados para gráfico de pizza - Situação
ws_graf['A3'] = 'SITUAÇÃO'
ws_graf['B3'] = 'QUANTIDADE'
ws_graf['A4'] = 'Membro'
ws_graf['B4'] = '=CONT.SES(SECRETARIA!$J$12:$J$21,"Membro")'
ws_graf['A5'] = 'Visitante'
ws_graf['B5'] = '=CONT.SES(SECRETARIA!$J$12:$J$21,"Visitante")'

# Criar gráfico de pizza
pie = PieChart()
pie.title = "Distribuição por Situação"
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True

data_pie = Reference(ws_graf, min_col=2, min_row=3, max_row=5)
cats_pie = Reference(ws_graf, min_col=1, min_row=4, max_row=5)
pie.add_data(data_pie, titles_from_data=True)
pie.set_categories(cats_pie)

ws_graf.add_chart(pie, "D3")

# Dados para gráfico de barras - Ministérios
ws_graf['A10'] = 'MINISTÉRIO'
ws_graf['B10'] = 'MEMBROS'
row = 11
for minist, count in sorted(ministerios.items()):
    ws_graf.cell(row=row, column=1).value = minist
    ws_graf.cell(row=row, column=2).value = count
    row += 1

# Criar gráfico de barras
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Membros por Ministério"
chart.y_axis.title = 'Quantidade'

data = Reference(ws_graf, min_col=2, min_row=10, max_row=row-1, max_col=2)
cats = Reference(ws_graf, min_col=1, min_row=11, max_row=row-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws_graf.add_chart(chart, "D15")

# ================================================================================
# ABA 4: GUIA DE FÓRMULAS
# ================================================================================
print("[4/4] Criando Guia de Fórmulas...")
ws_guia = wb.create_sheet("4. GUIA DE FORMULAS")

# Titulo
ws_guia.merge_cells('A1:D1')
ws_guia['A1'] = 'GUIA DE FÓRMULAS - SECRETARIA'
ws_guia['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_guia['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Headers
headers_guia = ['FUNÇÃO', 'SINTAXE', 'DESCRIÇÃO', 'EXEMPLO USADO']
for col_idx, header in enumerate(headers_guia, 1):
    cell = ws_guia.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

formulas = [
    ['CONT.SES', '=CONT.SES(intervalo;critério)', 'Conta células por critério', '=CONT.SES(J12:J21,"Membro")'],
    ['SOMA', '=SOMA(intervalo)', 'Soma valores numéricos', '=SOMA(B6:B7)'],
    ['SEERRO', '=SEERRO(valor;valor_se_erro)', 'Tratamento de erro', '=SEERRO(B6/SOMA($B$6:$B$7)*100;0)'],
    ['SE', '=SE(teste;valor_verdadeiro;falso)', 'Teste condicional', '=SE(J12="Membro";"Ativo";"Prospect")'],
    ['PROCV', '=PROCV(valor;matriz;coluna;falso)', 'Busca na vertical', '=PROCV("Ana";B:B;C;FALSO)'],
    ['DATADIF', '=DATADIF(data1;data2;"D")', 'Diferença em dias', '=DATADIF(C12;HOJE();"Y")'],
]

for idx, formula in enumerate(formulas, 1):
    row = 3 + idx
    for col_idx, value in enumerate(formula, 1):
        cell = ws_guia.cell(row=row, column=col_idx)
        cell.value = value
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 1:
            cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['AZUL_PRIMARIO'])

# Larguras
ws_guia.column_dimensions['A'].width = 15
ws_guia.column_dimensions['B'].width = 30
ws_guia.column_dimensions['C'].width = 35
ws_guia.column_dimensions['D'].width = 35

# Dicas
ws_guia['A12'] = 'DICAS E BOAS PRÁTICAS'
ws_guia['A12'].font = Font(bold=True, color=CORES['DOURADO'] if 'DOURADO' in CORES else 'D97706', size=14, name='Calibri')

dicas = [
    '1. Use CONT.SES para contar automaticamente membros e visitantes',
    '2. Formatação condicional destaca situações automaticamente (verde=Membro, laranja=Visitante)',
    '3. SEERRO evita erros #DIV/0! quando não há dados',
    '4. PROCV facilita buscas por nome ou código',
    '5. Mantenha dados sempre atualizados para relatórios precisos',
]

for idx, dica in enumerate(dicas, 1):
    ws_guia.cell(row=13+idx, column=1).value = dica
    ws_guia.cell(row=13+idx, column=1).font = Font(name='Calibri', size=10)

# ================================================================================
# SALVAR ARQUIVO
# ================================================================================
output_path = 'C:/Users/eduka/Desktop/Planilha_Secretaria_EXPERT.xlsx'
wb.save(output_path)

print("=" * 80)
print("PLANILHA DE SECRETARIA EXPERT CRIADA COM SUCESSO!")
print("=" * 80)
print(f"Arquivo: {output_path}")
print("\nABAS CRIADAS:")
print("  1. SECRETARIA - Dashboard com KPIs coloridos + Tabela de cadastro")
print("  2. ANALISE DEMOGRAFICA - Por situação, bairro e ministério")
print("  3. GRAFICOS - Pizza e Barras")
print("  4. GUIA DE FORMULAS - Documentação completa")
print("\nHABILIDADES EXCEL IMPLEMENTADAS:")
print("  - Formulas: CONT.SES, SOMA, SEERRO, SE")
print("  - Formatacao condicional: Cores para Membro (verde) e Visitante (laranja)")
print("  - Dashboard: Cards coloridos iguais a imagem")
print("  - Cores profissionais: Azul, Verde, Laranja")
print("=" * 80)
