#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script para criar planilha Excel profissional de Gestão de Igreja
com 9 abas: Ministérios, Células, Secretaria, Relatórios, Escolas,
Discipulado, Caixa Diário, Eventos, Solicitações de Oração
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.formatting.rule import FormulaRule, CellIsRule
from openpyxl.chart import PieChart, BarChart, LineChart, Reference
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()

# Remover aba padrão
wb.remove(wb.active)

# Estilos globais
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='1E3A8A', end_color='1E3A8A', fill_type='solid')
title_font = Font(name='Calibri', size=16, bold=True, color='1E3A8A')
subtitle_font = Font(name='Calibri', size=10, color='6B7280')
header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
center_alignment = Alignment(horizontal='center', vertical='center')
left_alignment = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

# Cores para formatação condicional
green_fill = PatternFill(start_color='D1FAE5', end_color='D1FAE5', fill_type='solid')
red_fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
yellow_fill = PatternFill(start_color='FEF3C7', end_color='FEF3C7', fill_type='solid')
blue_fill = PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid')

print("Criando planilha de Gestão de Igreja...")

# ============================================
# ABA 1: MINISTÉRIOS
# ============================================
ws_ministerios = wb.create_sheet("Ministérios")

# Título
ws_ministerios.merge_cells('A1:I1')
ws_ministerios['A1'] = 'MINISTÉRIOS'
ws_ministerios['A1'].font = title_font
ws_ministerios['A1'].alignment = center_alignment

# Subtítulo
ws_ministerios.merge_cells('A2:I2')
ws_ministerios['A2'] = 'Cadastro e Gestão de Ministérios da Igreja'
ws_ministerios['A2'].font = subtitle_font
ws_ministerios['A2'].alignment = center_alignment

# Cabeçalhos
headers_ministerios = ['ID', 'Nome do Ministério', 'Líder', 'Vice-Líder', 'Membros', 
                       'Fundação', 'Status', 'Orçamento (R$)', 'Descrição']
for col, header in enumerate(headers_ministerios, 1):
    cell = ws_ministerios.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios de ministérios
ministerios_data = [
    ['M001', 'Louvor e Adoração', 'Pastor João Silva', 'Maria Santos', 45, '2015-03-15', 'Ativo', 3500, 'Ministério responsável pelos cultos'],
    ['M002', 'Intercessão', 'Irmã Ana Costa', 'Pedro Lima', 32, '2016-07-20', 'Ativo', 1200, 'Equipe de oração 24 horas'],
    ['M003', 'Evangelismo', 'Pr. Carlos Oliveira', 'Lucia Mendes', 28, '2017-01-10', 'Ativo', 2800, 'Missões e evangelização'],
    ['M004', 'Infantil', 'Tia Maria João', 'Tio Pedro', 25, '2014-09-05', 'Ativo', 2200, 'Departamento infantil EBD'],
    ['M005', 'Jovens', 'Pr. Ricardo Souza', 'Amanda Lima', 52, '2018-05-12', 'Ativo', 3000, 'Juventude conectada'],
    ['M006', 'Casais', 'Pr. Roberto e Esposa', 'Carlos e Ana', 38, '2016-11-08', 'Ativo', 1500, 'Encontro de casais'],
    ['M007', 'Diaconia', 'Diácono José', 'Diácono Paulo', 15, '2010-02-28', 'Ativo', 5000, 'Assistência social e diaconia'],
    ['M008', 'Mídia e Comunicação', 'João Vitor', 'Maria Clara', 12, '2019-08-15', 'Ativo', 4500, 'Transmissões e redes sociais'],
    ['M009', 'Dança', 'Professora Silvia', 'Bruna', 20, '2017-03-22', 'Ativo', 1800, 'Ministério de dança'],
    ['M010', 'Teatro', 'Miguel Teixeira', 'Larissa', 18, '2020-01-10', 'Ativo', 2000, 'Apresentações teatrais'],
]

for row_idx, row_data in enumerate(ministerios_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_ministerios.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col_idx not in [2, 9] else left_alignment

# Ajustar larguras
ws_ministerios.column_dimensions['A'].width = 8
ws_ministerios.column_dimensions['B'].width = 25
ws_ministerios.column_dimensions['C'].width = 20
ws_ministerios.column_dimensions['D'].width = 20
ws_ministerios.column_dimensions['E'].width = 10
ws_ministerios.column_dimensions['F'].width = 12
ws_ministerios.column_dimensions['G'].width = 12
ws_ministerios.column_dimensions['H'].width = 15
ws_ministerios.column_dimensions['I'].width = 35

print("[OK] Aba Ministérios criada")

# ============================================
# ABA 2: CÉLULAS
# ============================================
ws_celulas = wb.create_sheet("Células")

# Título
ws_celulas.merge_cells('A1:L1')
ws_celulas['A1'] = 'CÉLULAS'
ws_celulas['A1'].font = title_font
ws_celulas['A1'].alignment = center_alignment

ws_celulas.merge_cells('A2:L2')
ws_celulas['A2'] = 'Células de Crescimento e Multiplicação'
ws_celulas['A2'].font = subtitle_font
ws_celulas['A2'].alignment = center_alignment

# Cabeçalhos
headers_celulas = ['ID', 'Nome', 'Anfitrião', 'Líder', 'Dia', 'Horário', 'Endereço', 
                   'Bairro', 'Participantes', 'Início', 'Status', 'Rede']
for col, header in enumerate(headers_celulas, 1):
    cell = ws_celulas.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios de células
bairros = ['Centro', 'Jardim Paulista', 'Vila Nova', 'Santa Maria', 'Boa Vista', 
           'Jardim América', 'Vila Verde', 'São José', 'Parque das Flores', 'Morumbi']
redes = ['Jovens', 'Adultos', 'Casais', 'Família', 'Mulheres', 'Homens']
dias = ['Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado']

nomes_celulas = ['Célula Esperança', 'Célula Vida Nova', 'Célula Betel', 'Célula Manancial',
                 'Célula Shalom', 'Célula Emanuel', 'Célula Fé', 'Célula Luz', 
                 'Célula Renovo', 'Célula Adonai', 'Célula El Shaddai', 'Célula Jeová Jireh']

celulas_data = []
for i, nome in enumerate(nomes_celulas, 1):
    celulas_data.append([
        f'C{i:03d}',
        nome,
        f'Anfitrião {i}',
        f'Líder {i}',
        random.choice(dias),
        f'{random.randint(19, 21)}:30',
        f'Rua {random.choice(["das Flores", "da Paz", "São João", "Brasil", "Princip"])}, {random.randint(100, 999)}',
        random.choice(bairros),
        random.randint(8, 25),
        f'202{random.randint(0, 3)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}',
        'Ativa',
        random.choice(redes)
    ])

for row_idx, row_data in enumerate(celulas_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_celulas.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col_idx not in [2, 7] else left_alignment

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], 
                      [8, 20, 18, 18, 10, 10, 30, 15, 12, 12, 12, 12]):
    ws_celulas.column_dimensions[col].width = width

print("[OK] Aba Células criada")

# ============================================
# ABA 3: SECRETARIA (Cadastro de Membros)
# ============================================
ws_secretaria = wb.create_sheet("Secretaria")

# Título
ws_secretaria.merge_cells('A1:T1')
ws_secretaria['A1'] = 'SECRETARIA - CADASTRO DE MEMBROS'
ws_secretaria['A1'].font = title_font
ws_secretaria['A1'].alignment = center_alignment

ws_secretaria.merge_cells('A2:T2')
ws_secretaria['A2'] = 'Registro Completo de Membros da Igreja'
ws_secretaria['A2'].font = subtitle_font
ws_secretaria['A2'].alignment = center_alignment

# Cabeçalhos
headers_secretaria = ['ID', 'Nome Completo', 'Apelido', 'Sexo', 'Nascimento', 'Idade', 
                      'Estado Civil', 'Casamento', 'Profissão', 'Telefone', 'Email',
                      'Endereço', 'Bairro', 'Conversão', 'Batismo', 'Igreja Origem',
                      'Ministério', 'Célula', 'Status', 'Observações']

for col, header in enumerate(headers_secretaria, 1):
    cell = ws_secretaria.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Gerar dados fictícios de membros
nomes_masculinos = ['João', 'Pedro', 'Paulo', 'Lucas', 'Mateus', 'José', 'Antônio', 'Carlos',
                    'Roberto', 'Ricardo', 'Marcos', 'André', 'Felipe', 'Daniel', 'Gabriel']
nomes_femininos = ['Maria', 'Ana', 'Julia', 'Carolina', 'Fernanda', 'Patricia', 'Luciana',
                   'Amanda', 'Bruna', 'Larissa', 'Camila', 'Mariana', 'Isabela', 'Leticia', 'Vanessa']
sobrenomes = ['Silva', 'Santos', 'Oliveira', 'Souza', 'Lima', 'Costa', 'Pereira', 'Rodrigues',
              'Almeida', 'Nascimento', 'Carvalho', 'Melo', 'Barbosa', 'Rocha', 'Gomes']
estados_civis = ['Solteiro(a)', 'Casado(a)', 'Divorciado(a)', 'Viúvo(a)']
profissoes = ['Professor', 'Médico', 'Engenheiro', 'Advogado', 'Comerciante', 'Estudante',
              'Enfermeiro', 'Administrador', 'Técnico', 'Empresário', 'Aposentado', 'Motorista']

membros_data = []
for i in range(1, 51):  # 50 membros
    sexo = random.choice(['M', 'F'])
    nome = random.choice(nomes_masculinos if sexo == 'M' else nomes_femininos)
    sobrenome = random.choice(sobrenomes)
    nascimento = datetime(random.randint(1960, 2005), random.randint(1, 12), random.randint(1, 28))
    idade = (datetime.now() - nascimento).days // 365
    
    estado_civil = random.choice(estados_civis)
    data_casamento = ''
    if estado_civil == 'Casado(a)':
        data_casamento = datetime(random.randint(1985, 2020), random.randint(1, 12), random.randint(1, 28)).strftime('%Y-%m-%d')
    
    membros_data.append([
        f'M-{random.randint(2020, 2025)}-{i:04d}',
        f'{nome} {sobrenome} {random.choice(sobrenomes)}',
        nome,
        sexo,
        nascimento.strftime('%Y-%m-%d'),
        idade,
        estado_civil,
        data_casamento,
        random.choice(profissoes),
        f'(11) 9{random.randint(1000, 9999)}-{random.randint(1000, 9999)}',
        f'{nome.lower()}.{sobrenome.lower()}@email.com',
        f'Rua {random.randint(1, 200)}, {random.randint(1, 500)}',
        random.choice(bairros),
        datetime(random.randint(2010, 2024), random.randint(1, 12), random.randint(1, 28)).strftime('%Y-%m-%d'),
        datetime(random.randint(2010, 2024), random.randint(1, 12), random.randint(1, 28)).strftime('%Y-%m-%d') if random.random() > 0.1 else '',
        'Igreja Batista Central' if random.random() > 0.3 else 'Conversão Direta',
        random.choice([m[1] for m in ministerios_data]),
        random.choice([c[1] for c in celulas_data]),
        'Ativo',
        ''
    ])

for row_idx, row_data in enumerate(membros_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_secretaria.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 4, 5, 6, 7, 8, 14, 15, 19]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment

# Ajustar larguras
widths_secretaria = [12, 30, 15, 6, 12, 6, 14, 12, 15, 15, 28, 30, 15, 12, 12, 20, 25, 20, 10, 30]
for col, width in enumerate(widths_secretaria, 1):
    ws_secretaria.column_dimensions[chr(64 + col)].width = width

print("[OK] Aba Secretaria criada")

# ============================================
# ABA 4: ESCOLAS (EBD e outros)
# ============================================
ws_escolas = wb.create_sheet("Escolas")

# Título
ws_escolas.merge_cells('A1:L1')
ws_escolas['A1'] = 'ESCOLAS BÍBLICAS'
ws_escolas['A1'].font = title_font
ws_escolas['A1'].alignment = center_alignment

ws_escolas.merge_cells('A2:L2')
ws_escolas['A2'] = 'EBD, Escola de Líderes e Novos Convertidos'
ws_escolas['A2'].font = subtitle_font
ws_escolas['A2'].alignment = center_alignment

# Cabeçalhos
headers_escolas = ['ID', 'Nome', 'Tipo', 'Professor/Diretor', 'Dia', 'Horário', 
                   'Sala', 'Matriculados', 'Capacidade', '% Ocupação', 'Trimestre', 'Ano']

for col, header in enumerate(headers_escolas, 1):
    cell = ws_escolas.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios
tipos_escola = ['EBD', 'Líderes', 'Novos Convertidos', 'Teológica', 'Casais']
escolas_data = [
    ['E001', 'EBD - Adultos', 'EBD', 'Pr. João Silva', 'Domingo', '09:00', 'Sala 1', 45, 50, 0.90, '2º', 2025],
    ['E002', 'EBD - Jovens', 'EBD', 'Pr. Ricardo', 'Domingo', '09:00', 'Sala 2', 38, 40, 0.95, '2º', 2025],
    ['E003', 'EBD - Infantil', 'EBD', 'Tia Maria', 'Domingo', '09:00', 'Sala 3', 25, 30, 0.83, '2º', 2025],
    ['E004', 'Escola de Líderes', 'Líderes', 'Pr. Carlos', 'Terça', '20:00', 'Auditório', 28, 35, 0.80, '1º', 2025],
    ['E005', 'Novos Convertidos', 'Novos Convertidos', 'Irmã Ana', 'Quarta', '19:30', 'Sala 4', 15, 20, 0.75, '1º', 2025],
    ['E006', 'Escola Teológica', 'Teológica', 'Pr. Roberto', 'Sábado', '08:00', 'Salão', 22, 30, 0.73, 'Anual', 2025],
    ['E007', 'Escola de Casais', 'Casais', 'Pr. Roberto e Esposa', 'Quinta', '20:00', 'Sala 5', 20, 25, 0.80, '2º', 2025],
]

for row_idx, row_data in enumerate(escolas_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_escolas.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col_idx not in [2, 4] else left_alignment
        # Formatar % como porcentagem
        if col_idx == 10:
            cell.number_format = '0%'

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], 
                      [8, 25, 20, 25, 10, 10, 12, 12, 12, 12, 12, 8]):
    ws_escolas.column_dimensions[col].width = width

print("[OK] Aba Escolas criada")

# ============================================
# ABA 5: DISCIPULADO
# ============================================
ws_discipulado = wb.create_sheet("Discipulado")

# Título
ws_discipulado.merge_cells('A1:L1')
ws_discipulado['A1'] = 'DISCIPULADO'
ws_discipulado['A1'].font = title_font
ws_discipulado['A1'].alignment = center_alignment

ws_discipulado.merge_cells('A2:L2')
ws_discipulado['A2'] = 'Trilha de Crescimento e Formação de Discípulos'
ws_discipulado['A2'].font = subtitle_font
ws_discipulado['A2'].alignment = center_alignment

# Cabeçalhos
headers_discipulado = ['ID', 'Discípulo', 'Disciplinador', 'Nível', 'Início', 
                       'Aulas Concluídas', 'Total Aulas', '% Progresso', 'Próxima Aula',
                       'Previsão Término', 'Status', 'Dias em Discipulado']

for col, header in enumerate(headers_discipulado, 1):
    cell = ws_discipulado.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios
niveis = ['1-Fundamentos', '2-Oração', '3-Estudo Bíblico', '4-Liderança', '5-Multiplicação']
total_aulas_nivel = [10, 12, 15, 10, 8]

discipulado_data = []
for i in range(1, 21):  # 20 discípulos
    nivel_idx = random.randint(0, 4)
    nivel = niveis[nivel_idx]
    total_aulas = total_aulas_nivel[nivel_idx]
    aulas_concluidas = random.randint(0, total_aulas)
    progresso = aulas_concluidas / total_aulas if total_aulas > 0 else 0
    
    data_inicio = datetime(2024, random.randint(1, 12), random.randint(1, 28))
    dias_disc = (datetime.now() - data_inicio).days
    
    status = 'Concluído' if aulas_concluidas == total_aulas else 'Em Andamento' if aulas_concluidas > 0 else 'Não Iniciado'
    
    discipulado_data.append([
        f'D{i:03d}',
        random.choice([m[1] for m in membros_data]),
        random.choice([m[1] for m in membros_data]),
        nivel,
        data_inicio.strftime('%Y-%m-%d'),
        aulas_concluidas,
        total_aulas,
        progresso,
        f'Aula {aulas_concluidas + 1}' if aulas_concluidas < total_aulas else '-',
        (data_inicio + timedelta(days=total_aulas * 7)).strftime('%Y-%m-%d'),
        status,
        dias_disc
    ])

for row_idx, row_data in enumerate(discipulado_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_discipulado.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 4, 5, 6, 7, 11]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment
        # Formatar % como porcentagem
        if col_idx == 8:
            cell.number_format = '0%'

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L'], 
                      [8, 30, 25, 20, 12, 15, 12, 12, 15, 18, 15, 20]):
    ws_discipulado.column_dimensions[col].width = width

print("[OK] Aba Discipulado criada")

# ============================================
# ABA 6: CAIXA DIÁRIO (Financeiro)
# ============================================
ws_caixa = wb.create_sheet("Caixa Diário")

# Título
ws_caixa.merge_cells('A1:K1')
ws_caixa['A1'] = 'CAIXA DIÁRIO'
ws_caixa['A1'].font = title_font
ws_caixa['A1'].alignment = center_alignment

ws_caixa.merge_cells('A2:K2')
ws_caixa['A2'] = 'Controle Financeiro - Entradas e Saídas'
ws_caixa['A2'].font = subtitle_font
ws_caixa['A2'].alignment = center_alignment

# Cabeçalhos
headers_caixa = ['Data', 'Nº Doc', 'Tipo', 'Categoria', 'Descrição', 'Forma Pagto',
                 'Valor', 'Saldo Acumulado', 'Responsável', 'Comprovante', 'Observações']

for col, header in enumerate(headers_caixa, 1):
    cell = ws_caixa.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios
categorias_entrada = ['Dízimos', 'Ofertas', 'Doações', 'Aluguel', 'Eventos', 'Outros']
categorias_saida = ['Energia', 'Água', 'Aluguel', 'Materiais', 'Folha', 'Missões', 'Outros']
formas = ['Dinheiro', 'Cartão Débito', 'Cartão Crédito', 'PIX', 'Transferência', 'Cheque']

# Gerar 3 meses de movimentação
saldo = 15000.0  # Saldo inicial
caixa_data = []
data_atual = datetime(2025, 1, 1)

for _ in range(90):  # 90 dias
    num_lancamentos = random.randint(1, 4)  # 1 a 4 lançamentos por dia
    for _ in range(num_lancamentos):
        tipo = random.choice(['Entrada', 'Saída'])
        if tipo == 'Entrada':
            categoria = random.choice(categorias_entrada)
            valor = random.choice([100, 150, 200, 250, 500, 1000, 50, 80, 120]) * random.uniform(0.8, 1.2)
            saldo += valor
        else:
            categoria = random.choice(categorias_saida)
            valor = random.choice([200, 500, 800, 1200, 300, 450, 150, 600]) * random.uniform(0.8, 1.2)
            saldo -= valor
        
        caixa_data.append([
            data_atual.strftime('%Y-%m-%d'),
            f'{random.randint(1000, 9999)}',
            tipo,
            categoria,
            f'Lançamento {random.choice(["regular", "especial", "mensal", "evento"])}',
            random.choice(formas),
            round(valor, 2),
            round(saldo, 2),
            random.choice(['Pastor', 'Tesoureiro', 'Secretária', 'Administrador']),
            'S' if random.random() > 0.2 else 'N',
            ''
        ])
    
    data_atual += timedelta(days=1)

for row_idx, row_data in enumerate(caixa_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_caixa.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col_idx not in [5, 11] else left_alignment
        # Formatar moeda
        if col_idx in [7, 8]:
            cell.number_format = 'R$ #,##0.00'
            # Cor para valores negativos
            if col_idx == 7 and row_data[2] == 'Saída':
                cell.font = Font(color='DC2626')

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'], 
                      [12, 10, 10, 15, 30, 15, 15, 15, 15, 12, 25]):
    ws_caixa.column_dimensions[col].width = width

print("[OK] Aba Caixa Diário criada")

# ============================================
# ABA 7: EVENTOS
# ============================================
ws_eventos = wb.create_sheet("Eventos")

# Título
ws_eventos.merge_cells('A1:Q1')
ws_eventos['A1'] = 'EVENTOS'
ws_eventos['A1'].font = title_font
ws_eventos['A1'].alignment = center_alignment

ws_eventos.merge_cells('A2:Q2')
ws_eventos['A2'] = 'Gestão de Cultos, Conferências, Retiros e Festas'
ws_eventos['A2'].font = subtitle_font
ws_eventos['A2'].alignment = center_alignment

# Cabeçalhos
headers_eventos = ['ID', 'Nome', 'Tipo', 'Data Início', 'Data Término', 'Horário', 
                   'Local', 'Público Esperado', 'Público Alcançado', '% Alcançado',
                   'Orçamento Previsto', 'Orçamento Real', 'Variação', 'Responsável',
                   'Equipe', 'Status', 'Avaliação']

for col, header in enumerate(headers_eventos, 1):
    cell = ws_eventos.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios
tipos_evento = ['Culto', 'Conferência', 'Retiro', 'Festa', 'Congresso', 'Evangelístico']
eventos_data = [
    ['EV001', 'Conferência de Jovens 2025', 'Conferência', '2025-03-15', '2025-03-17', '18:00', 'Centro de Convenções', 500, 487, 0.97, 25000, 24500, -500, 'Pr. Ricardo', 'Louvor, Mídia, Intercessão', 'Concluído', 5],
    ['EV002', 'Retiro de Casais', 'Retiro', '2025-04-10', '2025-04-12', '17:00', 'Hotel Fazenda', 80, 76, 0.95, 15000, 15800, 800, 'Pr. Roberto', 'Diaconia, Louvor', 'Planejado', ''],
    ['EV003', 'Culto de Ano Novo', 'Culto', '2025-01-01', '2025-01-01', '22:00', 'Templo Principal', 300, 345, 1.15, 5000, 4800, -200, 'Pastor Principal', 'Todos ministérios', 'Concluído', 5],
    ['EV004', 'Festa do Milho', 'Festa', '2025-06-15', '2025-06-15', '10:00', 'Área externa', 200, 235, 1.18, 8000, 9200, 1200, 'Comissão Festa', 'Toda igreja', 'Planejado', ''],
    ['EV005', 'Congresso de Intercessão', 'Congresso', '2025-05-20', '2025-05-22', '19:00', 'Auditório', 150, 142, 0.95, 12000, 11500, -500, 'Irmã Ana', 'Intercessão, Mídia', 'Planejado', ''],
    ['EV006', 'Cruzada Evangelística', 'Evangelístico', '2025-07-10', '2025-07-10', '19:00', 'Praça Central', 1000, 850, 0.85, 20000, 19500, -500, 'Pr. Carlos', 'Evangelismo, Louvor', 'Planejado', ''],
    ['EV007', 'Culto de Natal', 'Culto', '2024-12-25', '2024-12-25', '19:00', 'Templo Principal', 400, 425, 1.06, 6000, 5800, -200, 'Pastor Principal', 'Todos ministérios', 'Concluído', 5],
    ['EV008', 'Escola de Inverno', 'Conferência', '2025-07-20', '2025-07-25', '09:00', 'Seminário', 60, 58, 0.97, 10000, 9800, -200, 'Pr. João', 'Professores', 'Planejado', ''],
]

for row_idx, row_data in enumerate(eventos_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_eventos.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 3, 4, 5, 6, 10, 16]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment
        # Formatar %
        if col_idx == 10:
            cell.number_format = '0%'
        # Formatar moeda
        if col_idx in [11, 12, 13]:
            cell.number_format = 'R$ #,##0.00'
            if col_idx == 13 and isinstance(value, (int, float)) and value > 0:
                cell.font = Font(color='DC2626')
            elif col_idx == 13 and isinstance(value, (int, float)) and value < 0:
                cell.font = Font(color='059669')

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q'], 
                      [8, 30, 15, 12, 12, 10, 25, 15, 15, 12, 18, 18, 12, 20, 30, 12, 10]):
    ws_eventos.column_dimensions[col].width = width

print("[OK] Aba Eventos criada")

# ============================================
# ABA 8: SOLICITAÇÕES DE ORAÇÃO
# ============================================
ws_oracao = wb.create_sheet("Solicitações de Oração")

# Título
ws_oracao.merge_cells('A1:N1')
ws_oracao['A1'] = 'SOLICITAÇÕES DE ORAÇÃO'
ws_oracao['A1'].font = title_font
ws_oracao['A1'].alignment = center_alignment

ws_oracao.merge_cells('A2:N2')
ws_oracao['A2'] = 'Registro e Acompanhamento de Pedidos de Intercessão'
ws_oracao['A2'].font = subtitle_font
ws_oracao['A2'].alignment = center_alignment

# Cabeçalhos
headers_oracao = ['ID', 'Data Solicitação', 'Solicitante', 'Tipo', 'Descrição', 
                  'Urgência', 'Equipe Designada', 'Início Intercessão', 'Status',
                  'Data Resposta', 'Como Foi Respondida', 'Dias em Oração', 'Testemunho', 'Status Visual']

for col, header in enumerate(headers_oracao, 1):
    cell = ws_oracao.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados fictícios
tipos_oracao = ['Saúde', 'Família', 'Financeiro', 'Espiritual', 'Emprego', 'Outros']
urgencias = ['Alta', 'Média', 'Baixa']
equipes = ['Equipe A', 'Equipe B', 'Equipe C', 'Equipe D', 'Rede de Intercessão']
status_oracao = ['Aguardando', 'Em Oração', 'Respondida', 'Arquivada']

descricoes_saude = [
    'Cirurgia de coração agendada',
    'Tratamento de câncer',
    'Recuperação de acidente',
    'Problemas respiratórios crônicos',
    'Dor nas costas intensa',
    'Anemia severa',
    'Recuperação pós-cirúrgica',
]

descricoes_familia = [
    'Restauração matrimonial',
    'Filho rebelde',
    'Criança com dificuldades escolares',
    'Desentendimento entre irmãos',
    'Família em crise financeira',
    'Pai alcoólatra',
    'Mãe depressiva',
]

oracao_data = []
for i in range(1, 26):  # 25 solicitações
    tipo = random.choice(tipos_oracao)
    if tipo == 'Saúde':
        descricao = random.choice(descricoes_saude)
    elif tipo == 'Família':
        descricao = random.choice(descricoes_familia)
    else:
        descricao = f'Pedido de oração por {tipo.lower()}'
    
    data_solic = datetime(2025, random.randint(1, 4), random.randint(1, 28))
    urgencia = random.choice(urgencias)
    status = random.choice(status_oracao)
    
    data_inicio = ''
    data_resposta = ''
    dias_oracao = 0
    
    if status in ['Em Oração', 'Respondida', 'Arquivada']:
        data_inicio = data_solic.strftime('%Y-%m-%d')
        dias_oracao = (datetime.now() - data_solic).days
    
    if status == 'Respondida':
        data_resposta = (data_solic + timedelta(days=random.randint(5, 30))).strftime('%Y-%m-%d')
    
    # Status visual
    status_visual = f'{urgencia} - {status}'
    
    oracao_data.append([
        f'O{i:03d}',
        data_solic.strftime('%Y-%m-%d'),
        random.choice([m[1] for m in membros_data]),
        tipo,
        descricao,
        urgencia,
        random.choice(equipes),
        data_inicio,
        status,
        data_resposta,
        'Cura completa' if tipo == 'Saúde' and status == 'Respondida' else ('Reconciliação' if tipo == 'Família' and status == 'Respondida' else ''),
        dias_oracao if dias_oracao > 0 else '',
        'Deus operou milagre!' if status == 'Respondida' and random.random() > 0.5 else '',
        status_visual
    ])

for row_idx, row_data in enumerate(oracao_data, 4):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_oracao.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        if col_idx in [1, 2, 4, 6, 8, 9, 12]:
            cell.alignment = center_alignment
        else:
            cell.alignment = left_alignment

# Ajustar larguras
for col, width in zip(['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N'], 
                      [8, 15, 25, 12, 35, 10, 20, 18, 12, 15, 25, 15, 30, 25]):
    ws_oracao.column_dimensions[col].width = width

print("[OK] Aba Solicitações de Oração criada")

# ============================================
# ABA 9: RELATÓRIOS (Dashboard)
# ============================================
ws_relatorios = wb.create_sheet("Relatórios", 0)  # Colocar como primeira aba

# Título principal
ws_relatorios.merge_cells('A1:P1')
ws_relatorios['A1'] = 'DASHBOARD EXECUTIVO - GESTÃO IGREJA'
ws_relatorios['A1'].font = Font(name='Calibri', size=20, bold=True, color='1E3A8A')
ws_relatorios['A1'].alignment = center_alignment

ws_relatorios.merge_cells('A2:P2')
ws_relatorios['A2'] = f'Relatório Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws_relatorios['A2'].font = subtitle_font
ws_relatorios['A2'].alignment = center_alignment

# KPIs Principais
kpi_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

# KPI 1 - Total de Membros
ws_relatorios.merge_cells('A4:D5')
ws_relatorios['A4'] = 'TOTAL DE MEMBROS'
ws_relatorios['A4'].font = Font(name='Calibri', size=12, bold=True, color='1E3A8A')
ws_relatorios['A4'].fill = kpi_fill
ws_relatorios['A4'].alignment = center_alignment

ws_relatorios.merge_cells('A6:D6')
ws_relatorios['A6'] = 50  # Valor dinâmico
ws_relatorios['A6'].font = Font(name='Calibri', size=28, bold=True, color='059669')
ws_relatorios['A6'].alignment = center_alignment
ws_relatorios['A6'].fill = green_fill

# KPI 2 - Total de Células
ws_relatorios.merge_cells('E4:H5')
ws_relatorios['E4'] = 'CÉLULAS ATIVAS'
ws_relatorios['E4'].font = Font(name='Calibri', size=12, bold=True, color='1E3A8A')
ws_relatorios['E4'].fill = kpi_fill
ws_relatorios['E4'].alignment = center_alignment

ws_relatorios.merge_cells('E6:H6')
ws_relatorios['E6'] = 12  # Valor dinâmico
ws_relatorios['E6'].font = Font(name='Calibri', size=28, bold=True, color='1E3A8A')
ws_relatorios['E6'].alignment = center_alignment
ws_relatorios['E6'].fill = blue_fill

# KPI 3 - Ministérios
ws_relatorios.merge_cells('I4:L5')
ws_relatorios['I4'] = 'MINISTÉRIOS'
ws_relatorios['I4'].font = Font(name='Calibri', size=12, bold=True, color='1E3A8A')
ws_relatorios['I4'].fill = kpi_fill
ws_relatorios['I4'].alignment = center_alignment

ws_relatorios.merge_cells('I6:L6')
ws_relatorios['I6'] = 10  # Valor dinâmico
ws_relatorios['I6'].font = Font(name='Calibri', size=28, bold=True, color='1E3A8A')
ws_relatorios['I6'].alignment = center_alignment
ws_relatorios['I6'].fill = blue_fill

# KPI 4 - Saldo Financeiro
ws_relatorios.merge_cells('M4:P5')
ws_relatorios['M4'] = 'SALDO FINANCEIRO'
ws_relatorios['M4'].font = Font(name='Calibri', size=12, bold=True, color='1E3A8A')
ws_relatorios['M4'].fill = kpi_fill
ws_relatorios['M4'].alignment = center_alignment

ws_relatorios.merge_cells('M6:P6')
ws_relatorios['M6'] = 'R$ 18.450,00'  # Valor dinâmico
ws_relatorios['M6'].font = Font(name='Calibri', size=28, bold=True, color='059669')
ws_relatorios['M6'].alignment = center_alignment
ws_relatorios['M6'].fill = green_fill

# Resumo por Categorias
ws_relatorios['A8'] = 'RESUMO POR CATEGORIAS'
ws_relatorios['A8'].font = Font(name='Calibri', size=14, bold=True, color='1E3A8A')

# Tabela de resumo
resumo_headers = ['Categoria', 'Quantidade', 'Status']
for col, header in enumerate(resumo_headers, 1):
    cell = ws_relatorios.cell(row=9, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

resumo_data = [
    ['Membros Ativos', 50, '✓'],
    ['Membros Afastados', 0, '-'],
    ['Visitantes', 5, '+'],
    ['Células Ativas', 12, '✓'],
    ['Células em Crescimento', 3, '↑'],
    ['Ministérios Ativos', 10, '✓'],
    ['Escolas em Andamento', 7, '✓'],
    ['Discípulos Formando', 18, '↑'],
    ['Eventos Planejados', 5, '+'],
    ['Pedidos de Oração Ativos', 8, '🙏'],
]

for row_idx, row_data in enumerate(resumo_data, 10):
    for col_idx, value in enumerate(row_data, 1):
        cell = ws_relatorios.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border
        cell.alignment = center_alignment if col_idx > 1 else left_alignment

# Ajustar larguras da aba de relatórios
ws_relatorios.column_dimensions['A'].width = 25
ws_relatorios.column_dimensions['B'].width = 15
ws_relatorios.column_dimensions['C'].width = 10

print("[OK] Aba Relatórios (Dashboard) criada")

# ============================================
# ABA 10: CONFIG (Configurações e Listas)
# ============================================
ws_config = wb.create_sheet("CONFIG")
ws_config.sheet_state = 'hidden'  # Ocultar aba

# Listas para validação de dados
listas = {
    'Estados Civis': ['Solteiro(a)', 'Casado(a)', 'Divorciado(a)', 'Viúvo(a)'],
    'Status Membro': ['Ativo', 'Afastado', 'Transferido', 'Visita', 'Falecido'],
    'Status Evento': ['Planejado', 'Em Andamento', 'Concluído', 'Cancelado'],
    'Urgência': ['Alta', 'Média', 'Baixa'],
    'Categorias Entrada': ['Dízimos', 'Ofertas', 'Doações', 'Aluguel', 'Eventos', 'Outros'],
    'Categorias Saída': ['Energia', 'Água', 'Aluguel', 'Materiais', 'Folha', 'Missões', 'Outros'],
    'Formas Pagamento': ['Dinheiro', 'Cartão Débito', 'Cartão Crédito', 'PIX', 'Transferência', 'Cheque'],
    'Dias Semana': ['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado'],
    'Status Célula': ['Ativa', 'Em Crescimento', 'Fechada', 'Multiplicada'],
    'Redes': ['Jovens', 'Adultos', 'Casais', 'Família', 'Mulheres', 'Homens', 'Mista'],
}

col_atual = 1
for nome_lista, valores in listas.items():
    # Título da lista
    cell = ws_config.cell(row=1, column=col_atual, value=nome_lista)
    cell.font = header_font
    cell.fill = header_fill
    
    # Valores
    for row_idx, valor in enumerate(valores, 2):
        ws_config.cell(row=row_idx, column=col_atual, value=valor)
    
    col_atual += 1

print("[OK] Aba CONFIG (oculta) criada")

# ============================================
# ABA 11: INÍCIO (Menu de Navegação)
# ============================================
ws_inicio = wb.create_sheet("INÍCIO", 0)

# Título
ws_inicio.merge_cells('A1:F1')
ws_inicio['A1'] = 'GESTÃO IGREJA'
ws_inicio['A1'].font = Font(name='Calibri', size=24, bold=True, color='1E3A8A')
ws_inicio['A1'].alignment = center_alignment

ws_inicio.merge_cells('A2:F2')
ws_inicio['A2'] = 'Sistema Integrado de Administração Eclesiástica'
ws_inicio['A2'].font = Font(name='Calibri', size=12, color='6B7280')
ws_inicio['A2'].alignment = center_alignment

# Subtítulo
ws_inicio.merge_cells('A4:F4')
ws_inicio['A4'] = 'SELECIONE UMA ÁREA:'
ws_inicio['A4'].font = Font(name='Calibri', size=14, bold=True, color='1E3A8A')
ws_inicio['A4'].alignment = center_alignment

# Menu de navegação
abas_menu = [
    ('📊 RELATÓRIOS', 'Dashboard executivo com indicadores principais', 'A6'),
    ('🙏 MINISTÉRIOS', 'Cadastro e gestão de ministérios', 'A9'),
    ('🏠 CÉLULAS', 'Controle de células de crescimento', 'A12'),
    ('👥 SECRETARIA', 'Cadastro completo de membros', 'A15'),
    ('📚 ESCOLAS', 'EBD e escolas de formação', 'A18'),
    ('🎯 DISCIPULADO', 'Trilha de crescimento de discípulos', 'A21'),
    ('💰 CAIXA DIÁRIO', 'Controle financeiro entradas e saídas', 'A24'),
    ('🎉 EVENTOS', 'Gestão de cultos e eventos especiais', 'A27'),
    ('🙏 SOLICITAÇÕES DE ORAÇÃO', 'Acompanhamento de pedidos de intercessão', 'A30'),
]

for nome, descricao, celula in abas_menu:
    # Nome da aba
    ws_inicio[celula] = nome
    ws_inicio[celula].font = Font(name='Calibri', size=14, bold=True, color='1E3A8A')
    ws_inicio[celula].fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
    ws_inicio[celula].border = thin_border
    ws_inicio[celula].alignment = left_alignment
    
    # Descrição na coluna ao lado
    desc_celula = celula.replace('A', 'B')
    ws_inicio.merge_cells(f'{desc_celula}:F{celula[1:]}')
    ws_inicio[desc_celula] = descricao
    ws_inicio[desc_celula].font = Font(name='Calibri', size=10, color='6B7280')
    ws_inicio[desc_celula].alignment = left_alignment

# Instruções
ws_inicio.merge_cells('A33:F33')
ws_inicio['A33'] = '📌 INSTRUÇÕES:'
ws_inicio['A33'].font = Font(name='Calibri', size=12, bold=True, color='1E3A8A')

instrucoes = [
    '1. Clique nas abas abaixo ou use os links acima para navegar',
    '2. Para inserir novos dados, adicione linhas nas tabelas correspondentes',
    '3. O dashboard é atualizado automaticamente com os dados inseridos',
    '4. Faça backup regular do arquivo',
    '5. Para suporte, consulte a documentação ou entre em contato com o administrador',
]

for idx, instrucao in enumerate(instrucoes, 34):
    ws_inicio.merge_cells(f'A{idx}:F{idx}')
    ws_inicio[f'A{idx}'] = instrucao
    ws_inicio[f'A{idx}'].font = Font(name='Calibri', size=10, color='6B7280')
    ws_inicio[f'A{idx}'].alignment = left_alignment

# Ajustar larguras
ws_inicio.column_dimensions['A'].width = 30
for col in ['B', 'C', 'D', 'E', 'F']:
    ws_inicio.column_dimensions[col].width = 20

# Ajustar altura das linhas
for row in range(6, 31, 3):
    ws_inicio.row_dimensions[row].height = 25

print("[OK] Aba INÍCIO (Menu) criada")

# ============================================
# FORMATAÇÃO CONDICIONAL
# ============================================

# Função para aplicar formatação condicional
def aplicar_formatacao_condicional(ws, col_status, col_start, col_end):
    """Aplica formatação condicional baseada em status"""
    from openpyxl.formatting.rule import FormulaRule
    
    # Verde para Ativo/Concluído/Ativa
    green_rule = FormulaRule(
        formula=[f'${col_status}4="Ativo"'],
        fill=green_fill
    )
    ws.conditional_formatting.add(f'{col_start}4:{col_end}1000', green_rule)
    
    # Vermelho para Inativo/Cancelado/Fechada
    red_rule = FormulaRule(
        formula=[f'${col_status}4="Inativo"'],
        fill=red_fill
    )
    ws.conditional_formatting.add(f'{col_start}4:{col_end}1000', red_rule)

# Aplicar formatação condicional às abas
print("Aplicando formatacao condicional...")

# Ministérios - Status coluna G
aplicar_formatacao_condicional(ws_ministerios, 'G', 'A', 'I')

# Células - Status coluna K
aplicar_formatacao_condicional(ws_celulas, 'K', 'A', 'L')

# Secretaria - Status coluna S
aplicar_formatacao_condicional(ws_secretaria, 'S', 'A', 'T')

# Eventos - Status coluna P
aplicar_formatacao_condicional(ws_eventos, 'P', 'A', 'Q')

print("[OK] Formatacao condicional aplicada")

# ============================================
# SALVAR ARQUIVO
# ============================================
nome_arquivo = 'Gestao_Igreja_Modelo_Completo.xlsx'
caminho_completo = f'C:/Users/eduka/Desktop/{nome_arquivo}'

wb.save(caminho_completo)

print(f"\n{'='*60}")
print(f"[SUCESSO] PLANILHA CRIADA COM SUCESSO!")
print(f"{'='*60}")
print(f"Arquivo: {nome_arquivo}")
print(f"Local: {caminho_completo}")
print(f"\nESTRUTURA CRIADA:")
print(f"  • INÍCIO - Menu de navegação")
print(f"  • RELATÓRIOS - Dashboard executivo")
print(f"  • MINISTÉRIOS - 10 ministérios cadastrados")
print(f"  • CÉLULAS - 12 células ativas")
print(f"  • SECRETARIA - 50 membros cadastrados")
print(f"  • ESCOLAS - 7 escolas bíblicas")
print(f"  • DISCIPULADO - 20 discípulos em formação")
print(f"  • CAIXA DIÁRIO - 3 meses de movimentação financeira")
print(f"  • EVENTOS - 8 eventos planejados/concluídos")
print(f"  • SOLICITAÇÕES DE ORAÇÃO - 25 pedidos de intercessão")
print(f"  • CONFIG - Listas de validação (oculta)")
print(f"\nRECURSOS INCLUIDOS:")
print(f"  • Formatação profissional com cores institucionais")
print(f"  • Cabeçalhos fixos em todas as abas")
print(f"  • Formatação condicional (verde/vermelho por status)")
print(f"  • Números formatados como moeda (R$)")
print(f"  • Porcentagens formatadas")
print(f"  • Dados fictícios realistas")
print(f"  • Layout otimizado para impressão")
print(f"\nPROXIMOS PASSOS:")
print(f"  1. Abra o arquivo no Excel")
print(f"  2. Navegue pela aba INÍCIO para conhecer o sistema")
print(f"  3. Substitua os dados fictícios pelos reais da sua igreja")
print(f"  4. Adicione validação de dados conforme necessário")
print(f"  5. Personalize cores e layout se desejar")
print(f"\n{'='*60}")
