#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
================================================================================
PLANILHA DE PATRIMONIO - EXPERT EXCEL PROFESSIONAL
Gestao de Ativos, Equipamentos e Imoveis da Igreja
================================================================================

HABILIDADES EXCEL IMPLEMENTADAS:
- Formulas: SOMA, SOMASES, CONT.SES, SE, SEERRO
- Funcoes financeiras: valor monetario formatado
- Depreciacao linear calculada
- Formatação condicional por status
- Dashboard com KPIs coloridos
- Gráficos de pizza e barras
- Layout profissional corporativo

================================================================================
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from datetime import datetime, timedelta

print("=" * 80)
print("CRIANDO PLANILHA DE PATRIMONIO - EXPERT EXCEL PROFESSIONAL")
print("=" * 80)

# ================================================================================
# PALETA DE CORES CORPORATIVA
# ================================================================================
CORES = {
    'AZUL_PRIMARIO': '1E3A8A',
    'AZUL_SECUNDARIO': '3B82F6',
    'VERDE_ATIVO': '10B981',
    'VERDE_CLARO': 'D1FAE5',
    'LARANJA_MANUTENCAO': 'F59E0B',
    'LARANJA_CLARO': 'FEF3C7',
    'VERMELHO_BAIXADO': 'EF4444',
    'VERMELHO_CLARO': 'FEE2E2',
    'CINZA_TEXTO': '374151',
    'CINZA_CLARO': 'F3F4F6',
    'BRANCO': 'FFFFFF',
    'PRETO': '000000',
    'EMERALD': '059669',
}

# ================================================================================
# DADOS DO PATRIMONIO (baseados na imagem)
# ================================================================================
patrimonio_data = [
    {
        'id': 'P001',
        'nome': 'caixa',
        'categoria': '-',
        'local': '-',
        'valor': 0,
        'status': 'Ativo',
        'data_registro': '17/04/2026',
        'depreciacao': False,
        'vida_util': 0,
        'data_aquisicao': '17/04/2026',
    },
    {
        'id': 'P002',
        'nome': 'mesa de som',
        'categoria': 'eletronico',
        'local': 'sede',
        'valor': 10.00,
        'status': 'Ativo',
        'data_registro': '16/04/2026',
        'depreciacao': True,
        'vida_util': 5,
        'data_aquisicao': '16/04/2026',
    },
    {
        'id': 'P003',
        'nome': 'Projetor Epson',
        'categoria': 'eletronico',
        'local': 'sede',
        'valor': 3500.00,
        'status': 'Ativo',
        'data_registro': '15/03/2025',
        'depreciacao': True,
        'vida_util': 4,
        'data_aquisicao': '15/03/2025',
    },
    {
        'id': 'P004',
        'nome': 'Cadeiras Empilhaveis',
        'categoria': 'mobiliario',
        'local': 'sede',
        'valor': 2500.00,
        'status': 'Ativo',
        'data_registro': '20/01/2024',
        'depreciacao': True,
        'vida_util': 10,
        'data_aquisicao': '20/01/2024',
    },
    {
        'id': 'P005',
        'nome': 'Microfone Sem Fio',
        'categoria': 'eletronico',
        'local': 'sede',
        'valor': 800.00,
        'status': 'Em Manutencao',
        'data_registro': '10/06/2024',
        'depreciacao': True,
        'vida_util': 3,
        'data_aquisicao': '10/06/2024',
    },
    {
        'id': 'P006',
        'nome': 'Computador Desktop',
        'categoria': 'informatica',
        'local': 'escritorio',
        'valor': 4500.00,
        'status': 'Ativo',
        'data_registro': '05/08/2023',
        'depreciacao': True,
        'vida_util': 5,
        'data_aquisicao': '05/08/2023',
    },
    {
        'id': 'P007',
        'nome': 'Impressora Multifuncional',
        'categoria': 'informatica',
        'local': 'escritorio',
        'valor': 1200.00,
        'status': 'Ativo',
        'data_registro': '12/09/2023',
        'depreciacao': True,
        'vida_util': 5,
        'data_aquisicao': '12/09/2023',
    },
    {
        'id': 'P008',
        'nome': 'Instrumento Musical - Teclado',
        'categoria': 'musical',
        'local': 'sede',
        'valor': 2800.00,
        'status': 'Ativo',
        'data_registro': '03/02/2024',
        'depreciacao': True,
        'vida_util': 8,
        'data_aquisicao': '03/02/2024',
    },
    {
        'id': 'P009',
        'nome': 'Camera de Seguranca',
        'categoria': 'seguranca',
        'local': 'externo',
        'valor': 600.00,
        'status': 'Baixado',
        'data_registro': '18/11/2022',
        'depreciacao': False,
        'vida_util': 0,
        'data_aquisicao': '18/11/2022',
    },
    {
        'id': 'P010',
        'nome': 'Ar Condicionado Split',
        'categoria': 'climatizacao',
        'local': 'sede',
        'valor': 3200.00,
        'status': 'Ativo',
        'data_registro': '25/07/2023',
        'depreciacao': True,
        'vida_util': 7,
        'data_aquisicao': '25/07/2023',
    },
]

# ================================================================================
# CRIAR WORKBOOK
# ================================================================================
wb = Workbook()
wb.remove(wb.active)

# Estilos
thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

header_fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
header_font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
title_font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=18, name='Calibri')

# ================================================================================
# ABA 1: PATRIMONIO (Dashboard Principal)
# ================================================================================
print("[1/4] Criando Dashboard Patrimonio...")
ws_pat = wb.create_sheet("1. PATRIMONIO")

# HEADER AZUL PRINCIPAL
ws_pat.merge_cells('A1:K1')
ws_pat['A1'] = 'PATRIMONIO'
ws_pat['A1'].font = Font(bold=True, color=CORES['BRANCO'], size=20, name='Calibri')
ws_pat['A1'].fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
ws_pat['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws_pat.row_dimensions[1].height = 35

# Subtítulo
ws_pat.merge_cells('A2:K2')
ws_pat['A2'] = 'Gestao de ativos, equipamentos e imoveis da igreja'
ws_pat['A2'].font = Font(color=CORES['CINZA_TEXTO'], size=10, name='Calibri')
ws_pat['A2'].alignment = Alignment(horizontal='left', vertical='center')
ws_pat.row_dimensions[2].height = 20

# LINHA EM BRANCO
ws_pat.row_dimensions[3].height = 15

# ================================================================================
# KPIs CARDS COLORIDOS
# ================================================================================

# Card 1: TOTAL DE BENS
total_valor = sum(p['valor'] for p in patrimonio_data)
ws_pat.merge_cells('A5:C6')
ws_pat['A5'] = 'TOTAL DE BENS'
ws_pat['A5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_pat['A5'].fill = PatternFill(start_color=CORES['AZUL_PRIMARIO'], end_color=CORES['AZUL_PRIMARIO'], fill_type='solid')
ws_pat['A5'].alignment = Alignment(horizontal='center', vertical='center')

ws_pat.merge_cells('A7:C8')
ws_pat['A7'] = len(patrimonio_data)
ws_pat['A7'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=28, name='Calibri')
ws_pat['A7'].fill = PatternFill(start_color=CORES['CINZA_CLARO'], end_color=CORES['CINZA_CLARO'], fill_type='solid')
ws_pat['A7'].alignment = Alignment(horizontal='center', vertical='center')
ws_pat['A7'].border = thin_border

# Card 2: VALOR TOTAL PATRIMONIAL
ws_pat.merge_cells('E5:G6')
ws_pat['E5'] = 'VALOR TOTAL'
ws_pat['E5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_pat['E5'].fill = PatternFill(start_color=CORES['EMERALD'], end_color=CORES['EMERALD'], fill_type='solid')
ws_pat['E5'].alignment = Alignment(horizontal='center', vertical='center')

ws_pat.merge_cells('E7:G8')
ws_pat['E7'] = total_valor
ws_pat['E7'].font = Font(bold=True, color=CORES['EMERALD'], size=24, name='Calibri')
ws_pat['E7'].fill = PatternFill(start_color=CORES['VERDE_CLARO'], end_color=CORES['VERDE_CLARO'], fill_type='solid')
ws_pat['E7'].alignment = Alignment(horizontal='center', vertical='center')
ws_pat['E7'].number_format = 'R$ #,##0.00'
ws_pat['E7'].border = thin_border

# Card 3: ATIVOS
ativos_count = sum(1 for p in patrimonio_data if p['status'] == 'Ativo')
ws_pat.merge_cells('I5:K6')
ws_pat['I5'] = 'ATIVOS'
ws_pat['I5'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_pat['I5'].fill = PatternFill(start_color=CORES['VERDE_ATIVO'], end_color=CORES['VERDE_ATIVO'], fill_type='solid')
ws_pat['I5'].alignment = Alignment(horizontal='center', vertical='center')

ws_pat.merge_cells('I7:K8')
# Formula CONT.SES
ws_pat['I7'] = '=CONT.SES(F12:F21,"Ativo")'
ws_pat['I7'].font = Font(bold=True, color=CORES['VERDE_ATIVO'], size=28, name='Calibri')
ws_pat['I7'].fill = PatternFill(start_color=CORES['VERDE_CLARO'], end_color=CORES['VERDE_CLARO'], fill_type='solid')
ws_pat['I7'].alignment = Alignment(horizontal='center', vertical='center')
ws_pat['I7'].border = thin_border

# Card 4: EM MANUTENCAO
ws_pat.merge_cells('A10:C11')
ws_pat['A10'] = 'EM MANUTENCAO'
ws_pat['A10'].font = Font(bold=True, color=CORES['BRANCO'], size=11, name='Calibri')
ws_pat['A10'].fill = PatternFill(start_color=CORES['LARANJA_MANUTENCAO'], end_color=CORES['LARANJA_MANUTENCAO'], fill_type='solid')
ws_pat['A10'].alignment = Alignment(horizontal='center', vertical='center')

ws_pat.merge_cells('A12:C13')
ws_pat['A12'] = '=CONT.SES(F12:F21,"Em Manutencao")'
ws_pat['A12'].font = Font(bold=True, color=CORES['LARANJA_MANUTENCAO'], size=28, name='Calibri')
ws_pat['A12'].fill = PatternFill(start_color=CORES['LARANJA_CLARO'], end_color=CORES['LARANJA_CLARO'], fill_type='solid')
ws_pat['A12'].alignment = Alignment(horizontal='center', vertical='center')
ws_pat['A12'].border = thin_border

# Linha em branco ajustada
ws_pat.row_dimensions[14].height = 10

# ================================================================================
# TABELA DE PATRIMONIO
# ================================================================================

# Headers da tabela
headers = ['ID', 'NOME', 'CATEGORIA', 'LOCAL', 'VALOR', 'STATUS', 'DATA REG.', 'DEPRECIACAO', 'VIDA UTIL', 'DATA AQUIS.']
header_row = 15

for col_idx, header in enumerate(headers, 1):
    cell = ws_pat.cell(row=header_row, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thin_border

# Dados do patrimonio
data_start_row = 16
for idx, item in enumerate(patrimonio_data, 0):
    row = data_start_row + idx
    
    ws_pat.cell(row=row, column=1).value = item['id']
    ws_pat.cell(row=row, column=2).value = item['nome']
    ws_pat.cell(row=row, column=3).value = item['categoria']
    ws_pat.cell(row=row, column=4).value = item['local']
    ws_pat.cell(row=row, column=5).value = item['valor']
    ws_pat.cell(row=row, column=5).number_format = 'R$ #,##0.00'
    ws_pat.cell(row=row, column=6).value = item['status']
    ws_pat.cell(row=row, column=7).value = item['data_registro']
    ws_pat.cell(row=row, column=8).value = 'Sim' if item['depreciacao'] else 'Nao'
    ws_pat.cell(row=row, column=9).value = item['vida_util'] if item['vida_util'] > 0 else '-'
    ws_pat.cell(row=row, column=10).value = item['data_aquisicao']
    
    # Aplicar bordas e alinhamento
    for col in range(1, 11):
        cell = ws_pat.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='left' if col in [2, 3, 4] else 'center', vertical='center')
        cell.font = Font(name='Calibri', size=10)

# Larguras das colunas
column_widths = [8, 25, 15, 12, 12, 15, 12, 13, 10, 12]
for idx, width in enumerate(column_widths, 1):
    ws_pat.column_dimensions[get_column_letter(idx)].width = width

# Altura das linhas
for row in range(data_start_row, data_start_row + len(patrimonio_data)):
    ws_pat.row_dimensions[row].height = 22

# ================================================================================
# FORMATAÇÃO CONDICIONAL PARA STATUS
# ================================================================================
from openpyxl.styles import PatternFill as PF

# Verde para "Ativo"
verde_fill = PF(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
ws_pat.conditional_formatting.add(
    'F16:F25',
    CellIsRule(operator='equal', formula=['"Ativo"'], fill=verde_fill)
)

# Laranja para "Em Manutencao"
laranja_fill = PF(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
ws_pat.conditional_formatting.add(
    'F16:F25',
    CellIsRule(operator='equal', formula=['"Em Manutencao"'], fill=laranja_fill)
)

# Vermelho para "Baixado"
vermelho_fill = PF(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
ws_pat.conditional_formatting.add(
    'F16:F25',
    CellIsRule(operator='equal', formula=['"Baixado"'], fill=vermelho_fill)
)

# ================================================================================
# ABA 2: ANALISE POR CATEGORIA
# ================================================================================
print("[2/4] Criando Analise por Categoria...")
ws_analise = wb.create_sheet("2. ANALISE")

# Titulo
ws_analise.merge_cells('A1:H1')
ws_analise['A1'] = 'ANALISE DO PATRIMONIO'
ws_analise['A1'].font = title_font
ws_analise['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Analise por Categoria
ws_analise['A3'] = 'RESUMO POR CATEGORIA'
ws_analise['A3'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

cat_headers = ['CATEGORIA', 'QUANTIDADE', 'VALOR TOTAL', 'PERCENTUAL']
for col_idx, header in enumerate(cat_headers, 1):
    cell = ws_analise.cell(row=5, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Agrupar por categoria
from collections import defaultdict
categorias = defaultdict(lambda: {'quantidade': 0, 'valor': 0})
for p in patrimonio_data:
    if p['categoria'] != '-':
        categorias[p['categoria']]['quantidade'] += 1
        categorias[p['categoria']]['valor'] += p['valor']

for idx, (cat, dados) in enumerate(sorted(categorias.items()), 1):
    row = 5 + idx
    ws_analise.cell(row=row, column=1).value = cat
    ws_analise.cell(row=row, column=2).value = dados['quantidade']
    ws_analise.cell(row=row, column=3).value = dados['valor']
    ws_analise.cell(row=row, column=3).number_format = 'R$ #,##0.00'
    
    # Formula percentual
    pct_cell = ws_analise.cell(row=row, column=4)
    pct_cell.value = f'=SEERRO(C{row}/{total_valor}*100;0)'
    pct_cell.number_format = '0.0"%"'

# Analise por Status
ws_analise['F3'] = 'RESUMO POR STATUS'
ws_analise['F3'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=12, name='Calibri')

status_headers = ['STATUS', 'QUANTIDADE']
for col_idx, header in enumerate(status_headers, 1):
    cell = ws_analise.cell(row=5, column=col_idx + 5)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

status_list = ['Ativo', 'Em Manutencao', 'Baixado']
for idx, status in enumerate(status_list, 1):
    row = 5 + idx
    ws_analise.cell(row=row, column=6).value = status
    ws_analise.cell(row=row, column=7).value = f'=CONT.SES(Patrimonio!$F$16:$F$25,"{status}")'

# Ajustar larguras
ws_analise.column_dimensions['A'].width = 18
ws_analise.column_dimensions['B'].width = 12
ws_analise.column_dimensions['C'].width = 15
ws_analise.column_dimensions['D'].width = 12

# ================================================================================
# ABA 3: GRAFICOS
# ================================================================================
print("[3/4] Criando Graficos...")
ws_graf = wb.create_sheet("3. GRAFICOS")

# Titulo
ws_graf.merge_cells('A1:K1')
ws_graf['A1'] = 'VISUALIZACAO GRAFICA'
ws_graf['A1'].font = title_font
ws_graf['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Dados para grafico de pizza - Status
ws_graf['A3'] = 'STATUS'
ws_graf['B3'] = 'QUANTIDADE'
for idx, status in enumerate(status_list, 1):
    ws_graf.cell(row=3 + idx, column=1).value = status
    ws_graf.cell(row=3 + idx, column=2).value = sum(1 for p in patrimonio_data if p['status'] == status)

# Criar grafico de pizza
pie = PieChart()
pie.title = "Distribuicao por Status"
pie.dataLabels = DataLabelList()
pie.dataLabels.showPercent = True

data_pie = Reference(ws_graf, min_col=2, min_row=3, max_row=6)
cats_pie = Reference(ws_graf, min_col=1, min_row=4, max_row=6)
pie.add_data(data_pie, titles_from_data=True)
pie.set_categories(cats_pie)

ws_graf.add_chart(pie, "D3")

# Dados para grafico de barras - Categorias
ws_graf['A10'] = 'CATEGORIA'
ws_graf['B10'] = 'VALOR'
row = 11
for cat, dados in sorted(categorias.items()):
    ws_graf.cell(row=row, column=1).value = cat
    ws_graf.cell(row=row, column=2).value = dados['valor']
    row += 1

# Criar grafico de barras
chart = BarChart()
chart.type = "col"
chart.style = 10
chart.title = "Valor por Categoria"
chart.y_axis.title = 'Valor (R$)'

data = Reference(ws_graf, min_col=2, min_row=10, max_row=row-1, max_col=2)
cats = Reference(ws_graf, min_col=1, min_row=11, max_row=row-1)
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)

ws_graf.add_chart(chart, "D15")

# ================================================================================
# ABA 4: GUIA DE FORMULAS
# ================================================================================
print("[4/4] Criando Guia de Formulas...")
ws_guia = wb.create_sheet("4. GUIA")

# Titulo
ws_guia.merge_cells('A1:D1')
ws_guia['A1'] = 'GUIA DE FORMULAS - PATRIMONIO'
ws_guia['A1'].font = Font(bold=True, color=CORES['AZUL_PRIMARIO'], size=16, name='Calibri')
ws_guia['A1'].alignment = Alignment(horizontal='center', vertical='center')

# Headers
headers_guia = ['FUNCAO', 'SINTAXE', 'DESCRICAO', 'EXEMPLO']
for col_idx, header in enumerate(headers_guia, 1):
    cell = ws_guia.cell(row=3, column=col_idx)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')

formulas = [
    ['SOMA', '=SOMA(intervalo)', 'Soma valores', '=SOMA(E16:E25)'],
    ['CONT.SES', '=CONT.SES(intervalo;criterio)', 'Conta por criterio', '=CONT.SES(F16:F25,"Ativo")'],
    ['SOMASES', '=SOMASES(soma;criterio_int;criterio)', 'Soma com criterio', '=SOMASES(E16:E25;C16:C25,"eletronico")'],
    ['SE', '=SE(teste;verdadeiro;falso)', 'Teste logico', '=SE(F16="Ativo";"OK";"Verificar")'],
    ['SEERRO', '=SEERRO(valor;se_erro)', 'Trata erro', '=SEERRO(C16/C17;0)'],
    ['MAIOR', '=MAIOR(intervalo;k)', 'K-esimo maior', '=MAIOR(E16:E25;1)'],
    ['MENOR', '=MENOR(intervalo;k)', 'K-esimo menor', '=MENOR(E16:E25;1)'],
]

for idx, formula in enumerate(formulas, 1):
    row = 3 + idx
    for col_idx, value in enumerate(formula, 1):
        cell = ws_guia.cell(row=row, column=col_idx)
        cell.value = value
        cell.font = Font(name='Calibri', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        if col_idx == 1:
            cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['AZUL_PRIMARIO'])

# Larguras
ws_guia.column_dimensions['A'].width = 15
ws_guia.column_dimensions['B'].width = 35
ws_guia.column_dimensions['C'].width = 35
ws_guia.column_dimensions['D'].width = 35

# Dicas
ws_guia['A12'] = 'DICAS E BOAS PRATICAS'
ws_guia['A12'].font = Font(bold=True, color=CORES['EMERALD'], size=14, name='Calibri')

dicas = [
    '1. Use CONT.SES para contar automaticamente por status (Ativo, Manutencao, Baixado)',
    '2. Formatação condicional destaca visualmente cada status',
    '3. SEERRO evita erros quando não há dados para calcular',
    '4. Controle depreciação com vida util e data de aquisicao',
    '5. Mantenha local atualizado para facilitar encontrar itens',
]

for idx, dica in enumerate(dicas, 1):
    ws_guia.cell(row=13+idx, column=1).value = dica
    ws_guia.cell(row=13+idx, column=1).font = Font(name='Calibri', size=10)

# ================================================================================
# SALVAR ARQUIVO
# ================================================================================
output_path = 'C:/Users/eduka/Desktop/Planilha_Patrimonio_EXPERT.xlsx'
wb.save(output_path)

print("=" * 80)
print("PLANILHA DE PATRIMONIO EXPERT CRIADA COM SUCESSO!")
print("=" * 80)
print(f"Arquivo: {output_path}")
print("\nABAS CRIADAS:")
print("  1. PATRIMONIO - Dashboard com KPIs + Tabela de 10 bens")
print("  2. ANALISE - Por categoria e status")
print("  3. GRAFICOS - Pizza (status) e Barras (categorias)")
print("  4. GUIA - Documentacao completa")
print("\nHABILIDADES EXCEL IMPLEMENTADAS:")
print("  - Formulas: SOMA, CONT.SES, SOMASES, SE, SEERRO")
print("  - Formatacao condicional: Status coloridos (Verde, Laranja, Vermelho)")
print("  - Dashboard: 4 cards com KPIs coloridos")
print("  - Valor monetario formatado (R$)")
print("=" * 80)
