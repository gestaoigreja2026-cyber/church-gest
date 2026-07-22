#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Planilha Excel exclusiva para Gestão de Ministérios
Layout moderno com cards, dashboard e relatórios
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, GradientFill
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule, ColorScaleRule
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# Cores institucionais
AZUL_ESCURO = '1E3A8A'
AZUL_CLARO = '3B82F6'
AZUL_MEDIO = '60A5FA'
DOURADO = 'D4AF37'
VERDE = '059669'
VERMELHO = 'DC2626'
CINZA = '6B7280'
CINZA_CLARO = 'F3F4F6'
BRANCO = 'FFFFFF'

# Estilos
title_font = Font(name='Calibri', size=18, bold=True, color=AZUL_ESCURO)
card_title_font = Font(name='Calibri', size=14, bold=True, color=AZUL_ESCURO)
card_subtitle_font = Font(name='Calibri', size=10, color=CINZA)
header_font = Font(name='Calibri', size=11, bold=True, color=BRANCO)
header_fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type='solid')
blue_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
gold_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')
green_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')

thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB')
)

center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

print("Criando planilha exclusiva de Ministérios...")

# ============================================
# ABA 1: INICIO (Dashboard com Cards)
# ============================================
ws_inicio = wb.create_sheet("🏠 INÍCIO", 0)

# Cabeçalho principal
ws_inicio.merge_cells('A1:H1')
ws_inicio['A1'] = '🏛️ GESTÃO DE MINISTÉRIOS'
ws_inicio['A1'].font = Font(name='Calibri', size=22, bold=True, color=AZUL_ESCURO)
ws_inicio['A1'].alignment = center_alignment
ws_inicio.row_dimensions[1].height = 35

ws_inicio.merge_cells('A2:H2')
ws_inicio['A2'] = f'Dashboard atualizado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws_inicio['A2'].font = Font(name='Calibri', size=10, color=CINZA, italic=True)
ws_inicio['A2'].alignment = center_alignment
ws_inicio.row_dimensions[2].height = 20

# Linha em branco
ws_inicio.row_dimensions[3].height = 15

# CARDS DE RESUMO (Linha 4-7)
cards = [
    ('A4', 'C6', '⛪', 'Total de Ministérios', '12', AZUL_ESCURO, blue_fill),
    ('D4', 'F6', '👥', 'Membros Ativos', '387', VERDE, green_fill),
    ('G4', 'I6', '💰', 'Orçamento Total', 'R$ 26.300', DOURADO, gold_fill),
]

for col_start, col_end, emoji, label, value, color, fill in cards:
    start_row = int(col_start[1:])
    end_row = int(col_end[1:])
    start_col = ord(col_start[0]) - 64
    end_col = ord(col_end[0]) - 64
    
    # Preencher todo o card primeiro
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws_inicio.cell(row=row, column=col)
            cell.fill = fill
            cell.border = thin_border
    
    # Emoji (linha inicial, centralizado)
    cell_emoji = ws_inicio.cell(row=start_row, column=start_col)
    cell_emoji.value = emoji
    cell_emoji.font = Font(size=28)
    cell_emoji.alignment = Alignment(horizontal='center', vertical='center')
    
    # Label (linha do meio)
    cell_label = ws_inicio.cell(row=start_row + 1, column=start_col)
    cell_label.value = label
    cell_label.font = Font(name='Calibri', size=9, color=CINZA)
    cell_label.alignment = center_alignment
    
    # Valor (última linha)
    cell_value = ws_inicio.cell(row=end_row, column=start_col)
    cell_value.value = value
    cell_value.font = Font(name='Calibri', size=18, bold=True, color=color)
    cell_value.alignment = center_alignment
    
    # Mesclar células de cada linha do card
    for row in range(start_row, end_row + 1):
        ws_inicio.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)

# Linha em branco
ws_inicio.row_dimensions[8].height = 20

# TABELA DE MINISTÉRIOS EM CARDS
ws_inicio['A9'] = '📋 VISÃO GERAL DOS MINISTÉRIOS'
ws_inicio['A9'].font = Font(name='Calibri', size=14, bold=True, color=AZUL_ESCURO)
ws_inicio.merge_cells('A9:I9')

# Cabeçalhos da tabela
headers = ['Ícone', 'Ministério', 'Líder', 'Membros', 'Reuniões', 'Status', 'Orçamento', 'Ações']
for col, header in enumerate(headers, 1):
    cell = ws_inicio.cell(row=10, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Dados dos ministérios
ministerios = [
    ('🎨', 'Artes', 'Maria Silva', 15, 4, 'Ativo', 1800),
    ('💑', 'Casais', 'Pastor Roberto', 38, 6, 'Ativo', 1500),
    ('📚', 'Ensino', 'Professora Ana', 25, 8, 'Ativo', 2200),
    ('⚽', 'Esportes', 'Carlos Mendes', 20, 3, 'Ativo', 2000),
    ('🎵', 'Louvor', 'Pastor João', 45, 12, 'Ativo', 3500),
    ('🙏', 'Intercessão', 'Irmã Ana Costa', 32, 24, 'Ativo', 1200),
    ('📢', 'Evangelismo', 'Pr. Carlos', 28, 6, 'Ativo', 2800),
    ('👶', 'Infantil', 'Tia Maria João', 25, 8, 'Ativo', 2200),
    ('🎯', 'Jovens', 'Pr. Ricardo', 52, 10, 'Ativo', 3000),
    ('🏥', 'Diaconia', 'Diácono José', 15, 4, 'Ativo', 5000),
    ('📱', 'Mídia', 'João Vitor', 12, 6, 'Ativo', 4500),
    ('💃', 'Dança', 'Prof. Silvia', 20, 5, 'Ativo', 1800),
]

for row_idx, (icone, nome, lider, membros, reunioes, status, orcamento) in enumerate(ministerios, 11):
    ws_inicio.cell(row=row_idx, column=1, value=icone).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=2, value=nome).font = Font(bold=True, color=AZUL_ESCURO)
    ws_inicio.cell(row=row_idx, column=3, value=lider)
    ws_inicio.cell(row=row_idx, column=4, value=membros).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=5, value=reunioes).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=6, value=status).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=7, value=f'R$ {orcamento:,.0f}').alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=8, value='📊 Ver Relatório').alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=8).font = Font(color=AZUL_CLARO, underline='single')
    
    for col in range(1, 9):
        ws_inicio.cell(row=row_idx, column=col).border = thin_border
        if row_idx % 2 == 0:
            ws_inicio.cell(row=row_idx, column=col).fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

# Ajustar larguras
ws_inicio.column_dimensions['A'].width = 8
ws_inicio.column_dimensions['B'].width = 18
ws_inicio.column_dimensions['C'].width = 22
ws_inicio.column_dimensions['D'].width = 10
ws_inicio.column_dimensions['E'].width = 10
ws_inicio.column_dimensions['F'].width = 12
ws_inicio.column_dimensions['G'].width = 15
ws_inicio.column_dimensions['H'].width = 18

print("[OK] Aba INICIO criada")

# ============================================
# ABA 2: CADASTRO (Dados completos)
# ============================================
ws_cadastro = wb.create_sheet("📋 Cadastro")

# Título
ws_cadastro.merge_cells('A1:K1')
ws_cadastro['A1'] = '📋 CADASTRO COMPLETO DE MINISTÉRIOS'
ws_cadastro['A1'].font = title_font
ws_cadastro['A1'].alignment = center_alignment

ws_cadastro.merge_cells('A2:K2')
ws_cadastro['A2'] = 'Informações detalhadas de todos os ministérios'
ws_cadastro['A2'].font = Font(size=10, color=CINZA, italic=True)
ws_cadastro['A2'].alignment = center_alignment

# Cabeçalhos
headers_cad = ['Código', 'Ícone', 'Nome', 'Líder', 'Vice-Líder', 'Membros', 'Fundação', 
               'Status', 'Orçamento Mensal', 'Descrição', 'Objetivos']
for col, header in enumerate(headers_cad, 1):
    cell = ws_cadastro.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Dados detalhados
descricoes = [
    'Ministério de artes visuais, teatro e expressão criativa',
    'Encontros e atividades para casais da igreja',
    'Escola Bíblica Dominical e formação teológica',
    'Atividades esportivas e recreativas para a comunidade',
    'Louvor, adoração e música nos cultos',
    'Oração intercessória 24 horas pela igreja',
    'Evangelização, missões e novos convertidos',
    'Ministério infantil e crianças',
    'Juventude e adolescentes',
    'Assistência social e ajuda aos necessitados',
    'Transmissões, fotos, vídeos e redes sociais',
    'Dança, coreografias e apresentações',
]

objetivos = [
    'Desenvolver talentos artísticos para Deus',
    'Fortalecer o casamento e família',
    'Ensinar a Palavra de Deus de forma didática',
    'Promover saúde e comunhão através do esporte',
    'Levar a igreja à presença de Deus',
    'Interceder pela igreja, cidade e nações',
    'Ganhar almas para o Reino de Deus',
    'Ensinar as crianças nos caminhos do Senhor',
    'Discipular jovens para Cristo',
    'Praticar o amor ao próximo',
    'Comunicar a mensagem da igreja',
    'Adorar a Deus através da dança',
]

for row_idx, ((icone, nome, lider, membros, reunioes, status, orcamento), desc, obj) in enumerate(zip(ministerios, descricoes, objetivos), 5):
    ws_cadastro.cell(row=row_idx, column=1, value=f'M{row_idx-4:03d}').alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=2, value=icone).alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=3, value=nome).font = Font(bold=True)
    ws_cadastro.cell(row=row_idx, column=4, value=lider)
    ws_cadastro.cell(row=row_idx, column=5, value=f'Vice {nome}').font = Font(italic=True, color=CINZA)
    ws_cadastro.cell(row=row_idx, column=6, value=membros).alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=7, value=f'201{random.randint(0, 9)}-{random.randint(1, 12):02d}-15').alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=8, value=status).alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=9, value=orcamento).alignment = center_alignment
    ws_cadastro.cell(row=row_idx, column=9).number_format = 'R$ #,##0'
    ws_cadastro.cell(row=row_idx, column=10, value=desc).font = Font(size=9)
    ws_cadastro.cell(row=row_idx, column=11, value=obj).font = Font(size=9)
    
    for col in range(1, 12):
        cell = ws_cadastro.cell(row=row_idx, column=col)
        cell.border = thin_border
        if status == 'Ativo':
            cell.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')

# Ajustar larguras
widths_cad = [10, 8, 15, 22, 20, 10, 12, 10, 18, 40, 35]
for col, width in enumerate(widths_cad, 1):
    ws_cadastro.column_dimensions[get_column_letter(col)].width = width

print("[OK] Aba Cadastro criada")

# ============================================
# ABA 3: MEMBROS POR MINISTÉRIO
# ============================================
ws_membros = wb.create_sheet("👥 Membros")

ws_membros.merge_cells('A1:H1')
ws_membros['A1'] = '👥 MEMBROS POR MINISTÉRIO'
ws_membros['A1'].font = title_font
ws_membros['A1'].alignment = center_alignment

ws_membros.merge_cells('A2:H2')
ws_membros['A2'] = 'Distribuição de membros em cada ministério'
ws_membros['A2'].font = Font(size=10, color=CINZA, italic=True)
ws_membros['A2'].alignment = center_alignment

# Dados de membros fictícios por ministério
headers_mem = ['Ministério', 'Nome', 'Função', 'Entrada', 'Telefone', 'Email', 'Status', 'Observações']
for col, header in enumerate(headers_mem, 1):
    cell = ws_membros.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Gerar membros para cada ministério
row_atual = 5
for icone, nome_min, lider, total_membros, reunioes, status, orcamento in ministerios:
    funcoes = ['Membro', 'Coordenador', 'Secretário', 'Tesoureiro', 'Voluntário']
    
    for i in range(min(total_membros, 8)):  # Máx 8 membros por ministério na planilha
        ws_membros.cell(row=row_atual, column=1, value=f'{icone} {nome_min}').font = Font(bold=True)
        ws_membros.cell(row=row_atual, column=2, value=f'Membro {i+1} {nome_min}')
        ws_membros.cell(row=row_atual, column=3, value='Líder' if i == 0 else random.choice(funcoes)).alignment = center_alignment
        ws_membros.cell(row=row_atual, column=4, value=f'202{random.randint(2, 4)}-{random.randint(1, 12):02d}-{random.randint(1, 28):02d}').alignment = center_alignment
        ws_membros.cell(row=row_atual, column=5, value=f'(11) 9{random.randint(1000, 9999)}-{random.randint(1000, 9999)}').alignment = center_alignment
        ws_membros.cell(row=row_atual, column=6, value=f'membro{i+1}@{nome_min.lower().replace(" ", "")}.com')
        ws_membros.cell(row=row_atual, column=7, value='Ativo').alignment = center_alignment
        ws_membros.cell(row=row_atual, column=8, value='')
        
        for col in range(1, 9):
            ws_membros.cell(row=row_atual, column=col).border = thin_border
            if row_atual % 2 == 0:
                ws_membros.cell(row=row_atual, column=col).fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')
        
        row_atual += 1

# Ajustar larguras
widths_mem = [20, 25, 15, 12, 15, 30, 10, 25]
for col, width in enumerate(widths_mem, 1):
    ws_membros.column_dimensions[get_column_letter(col)].width = width

print("[OK] Aba Membros criada")

# ============================================
# ABA 4: FINANCEIRO
# ============================================
ws_financeiro = wb.create_sheet("💰 Financeiro")

ws_financeiro.merge_cells('A1:H1')
ws_financeiro['A1'] = '💰 CONTROLE FINANCEIRO POR MINISTÉRIO'
ws_financeiro['A1'].font = title_font
ws_financeiro['A1'].alignment = center_alignment

ws_financeiro.merge_cells('A2:H2')
ws_financeiro['A2'] = 'Orçamentos, despesas e análise financeira'
ws_financeiro['A2'].font = Font(size=10, color=CINZA, italic=True)
ws_financeiro['A2'].alignment = center_alignment

# Resumo financeiro
ws_financeiro['A4'] = 'RESUMO FINANCEIRO'
ws_financeiro['A4'].font = Font(bold=True, size=12, color=AZUL_ESCURO)

resumo_fin = [
    ('Orçamento Total Anual:', 315600, AZUL_ESCURO),
    ('Total Gasto no Mês:', 26100, VERMELHO),
    ('Saldo Disponível:', 289500, VERDE),
    ('Maior Orçamento:', 5000, DOURADO),
]

for idx, (label, valor, cor) in enumerate(resumo_fin, 5):
    ws_financeiro.cell(row=idx, column=1, value=label).font = Font(bold=True)
    cell_valor = ws_financeiro.cell(row=idx, column=2, value=valor)
    cell_valor.font = Font(bold=True, color=cor)
    cell_valor.number_format = 'R$ #,##0'
    ws_financeiro.merge_cells(f'C{idx}:D{idx}')

# Tabela detalhada
headers_fin = ['Ministério', 'Orçamento Mensal', 'Gasto Mês', 'Saldo', '% Executado', 'Status Orçamentário']
for col, header in enumerate(headers_fin, 1):
    cell = ws_financeiro.cell(row=10, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for row_idx, (icone, nome, lider, membros, reunioes, status, orcamento) in enumerate(ministerios, 11):
    gasto = orcamento * random.uniform(0.6, 0.95)
    saldo = orcamento - gasto
    pct = gasto / orcamento
    
    ws_financeiro.cell(row=row_idx, column=1, value=f'{icone} {nome}')
    ws_financeiro.cell(row=row_idx, column=2, value=orcamento).number_format = 'R$ #,##0'
    ws_financeiro.cell(row=row_idx, column=2).alignment = center_alignment
    ws_financeiro.cell(row=row_idx, column=3, value=gasto).number_format = 'R$ #,##0'
    ws_financeiro.cell(row=row_idx, column=3).alignment = center_alignment
    ws_financeiro.cell(row=row_idx, column=4, value=saldo).number_format = 'R$ #,##0'
    ws_financeiro.cell(row=row_idx, column=4).alignment = center_alignment
    ws_financeiro.cell(row=row_idx, column=5, value=pct).number_format = '0%'
    ws_financeiro.cell(row=row_idx, column=5).alignment = center_alignment
    
    status_orc = '✅ OK' if pct < 0.9 else '⚠️ Atenção' if pct < 1 else '🔴 Estourado'
    ws_financeiro.cell(row=row_idx, column=6, value=status_orc).alignment = center_alignment
    
    for col in range(1, 7):
        ws_financeiro.cell(row=row_idx, column=col).border = thin_border

# Ajustar larguras
widths_fin = [25, 18, 15, 15, 14, 22]
for col, width in enumerate(widths_fin, 1):
    ws_financeiro.column_dimensions[get_column_letter(col)].width = width

print("[OK] Aba Financeiro criada")

# ============================================
# ABA 5: RELATÓRIOS E ANÁLISES
# ============================================
ws_relatorios = wb.create_sheet("📊 Relatórios")

ws_relatorios.merge_cells('A1:H1')
ws_relatorios['A1'] = '📊 RELATÓRIOS E ANÁLISES'
ws_relatorios['A1'].font = title_font
ws_relatorios['A1'].alignment = center_alignment

ws_relatorios.merge_cells('A2:H2')
ws_relatorios['A2'] = 'Indicadores e métricas de desempenho'
ws_relatorios['A2'].font = Font(size=10, color=CINZA, italic=True)
ws_relatorios['A2'].alignment = center_alignment

# Análises
analises = [
    ('📈', 'Maior Ministério', 'Louvor', '45 membros', AZUL_ESCURO),
    ('📉', 'Menor Ministério', 'Mídia', '12 membros', CINZA),
    ('💰', 'Maior Orçamento', 'Diaconia', 'R$ 5.000/mês', DOURADO),
    ('⏰', 'Mais Reuniões', 'Intercessão', '24/mês', VERDE),
    ('🎯', 'Maior Crescimento', 'Jovens', '+15% no ano', VERDE),
    ('📊', 'Média de Membros', 'Todos', '32 membros', AZUL_CLARO),
]

row_atual = 4
for emoji, titulo, minist, valor, cor in analises:
    ws_relatorios.cell(row=row_atual, column=1, value=emoji).font = Font(size=20)
    ws_relatorios.cell(row=row_atual, column=1).alignment = center_alignment
    ws_relatorios.cell(row=row_atual, column=2, value=titulo).font = Font(bold=True, size=11)
    ws_relatorios.cell(row=row_atual, column=3, value=minist).font = Font(color=cor, bold=True)
    ws_relatorios.cell(row=row_atual, column=4, value=valor).font = Font(bold=True)
    ws_relatorios.cell(row=row_atual, column=4).alignment = center_alignment
    row_atual += 1

# Ranking
ws_relatorios[f'A{row_atual+1}'] = '🏆 RANKING DE MINISTÉRIOS'
ws_relatorios[f'A{row_atual+1}'].font = Font(bold=True, size=12, color=AZUL_ESCURO)
ws_relatorios.merge_cells(f'A{row_atual+1}:D{row_atual+1}')

headers_rank = ['Posição', 'Ministério', 'Pontuação', 'Destaque']
for col, header in enumerate(headers_rank, 1):
    cell = ws_relatorios.cell(row=row_atual+3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

ranking = [
    ('🥇', 'Louvor', 95, 'Maior equipe'),
    ('🥈', 'Jovens', 88, 'Maior crescimento'),
    ('🥉', 'Diaconia', 82, 'Maior impacto social'),
    ('4', 'Intercessão', 80, 'Maior dedicação'),
    ('5', 'Infantil', 78, 'Futuro da igreja'),
]

for idx, (pos, minis, pts, dest) in enumerate(ranking, row_atual+4):
    ws_relatorios.cell(row=idx, column=1, value=pos).alignment = center_alignment
    ws_relatorios.cell(row=idx, column=2, value=minis).font = Font(bold=True)
    ws_relatorios.cell(row=idx, column=3, value=pts).alignment = center_alignment
    ws_relatorios.cell(row=idx, column=4, value=dest).font = Font(italic=True, color=CINZA)
    
    for col in range(1, 5):
        ws_relatorios.cell(row=idx, column=col).border = thin_border

# Ajustar larguras
widths_rel = [12, 20, 12, 25]
for col, width in enumerate(widths_rel, 1):
    ws_relatorios.column_dimensions[get_column_letter(col)].width = width

print("[OK] Aba Relatórios criada")

# ============================================
# ABA 6: REUNIÕES E ATIVIDADES
# ============================================
ws_reunioes = wb.create_sheet("📅 Reuniões")

ws_reunioes.merge_cells('A1:H1')
ws_reunioes['A1'] = '📅 CONTROLE DE REUNIÕES'
ws_reunioes['A1'].font = title_font
ws_reunioes['A1'].alignment = center_alignment

headers_reun = ['Data', 'Ministério', 'Tipo', 'Tema', 'Presentes', 'Faltas', 'Observações', 'Próxima Ação']
for col, header in enumerate(headers_reun, 1):
    cell = ws_reunioes.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

# Gerar reuniões fictícias
tipos_reun = ['Planejamento', 'Treinamento', 'Culto', 'Evento', 'Integração']
temas = ['Planejamento Trimestral', 'Treinamento de Líderes', 'Culto de Missões', 
         'Workshop de Canto', 'Dinâmica de Grupo', 'Estudo Bíblico', 'Preparação de Evento']

for row_idx in range(5, 25):
    icone, nome_min, _, membros, _, _, _ = random.choice(ministerios)
    
    ws_reunioes.cell(row=row_idx, column=1, value=f'2025-{random.randint(1, 6):02d}-{random.randint(1, 28):02d}').alignment = center_alignment
    ws_reunioes.cell(row=row_idx, column=2, value=f'{icone} {nome_min}')
    ws_reunioes.cell(row=row_idx, column=3, value=random.choice(tipos_reun)).alignment = center_alignment
    ws_reunioes.cell(row=row_idx, column=4, value=random.choice(temas))
    presentes = random.randint(5, membros)
    ws_reunioes.cell(row=row_idx, column=5, value=presentes).alignment = center_alignment
    ws_reunioes.cell(row=row_idx, column=6, value=membros - presentes).alignment = center_alignment
    ws_reunioes.cell(row=row_idx, column=7, value='Reunião produtiva' if random.random() > 0.3 else 'Faltas justificadas')
    ws_reunioes.cell(row=row_idx, column=8, value='Agendar próxima' if random.random() > 0.5 else 'Follow up pendente')
    
    for col in range(1, 9):
        ws_reunioes.cell(row=row_idx, column=col).border = thin_border

# Ajustar larguras
widths_reun = [12, 20, 15, 30, 10, 10, 25, 25]
for col, width in enumerate(widths_reun, 1):
    ws_reunioes.column_dimensions[get_column_letter(col)].width = width

print("[OK] Aba Reuniões criada")

# ============================================
# SALVAR ARQUIVO
# ============================================
nome_arquivo = 'Planilha_Ministerios_Completa.xlsx'
caminho_completo = f'C:/Users/eduka/Desktop/{nome_arquivo}'

wb.save(caminho_completo)

print(f"\n{'='*60}")
print(f"[SUCESSO] PLANILHA DE MINISTERIOS CRIADA!")
print(f"{'='*60}")
print(f"Arquivo: {nome_arquivo}")
print(f"Local: {caminho_completo}")
print(f"\nABAS CRIADAS:")
print(f"  • 🏠 INÍCIO - Dashboard com cards de ministérios")
print(f"  • 📋 Cadastro - Dados completos de todos os ministérios")
print(f"  • 👥 Membros - Lista de membros por ministério")
print(f"  • 💰 Financeiro - Controle orçamentário detalhado")
print(f"  • 📊 Relatórios - Análises e ranking")
print(f"  • 📅 Reuniões - Controle de atividades")
print(f"\nRECURSOS:")
print(f"  • Icones emojis em todas as abas")
print(f"  • Cards visuais na aba inicial")
print(f"  • Tabelas formatadas com cores")
print(f"  • Dashboard com indicadores")
print(f"  • 12 ministérios cadastrados")
print(f"  • Dados financeiros realistas")
print(f"\n{'='*60}")
