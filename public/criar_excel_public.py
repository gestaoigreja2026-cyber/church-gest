#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Criar planilha Excel na pasta public para download
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime, timedelta
import random

# Criar workbook
wb = Workbook()
wb.remove(wb.active)

# Cores
AZUL_ESCURO = '1E3A8A'
AZUL_CLARO = '3B82F6'
DOURADO = 'D4AF37'
VERDE = '059669'
VERMELHO = 'DC2626'
CINZA = '6B7280'

# Estilos
title_font = Font(name='Calibri', size=18, bold=True, color=AZUL_ESCURO)
header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
header_fill = PatternFill(start_color=AZUL_ESCURO, end_color=AZUL_ESCURO, fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='E5E7EB'),
    right=Side(style='thin', color='E5E7EB'),
    top=Side(style='thin', color='E5E7EB'),
    bottom=Side(style='thin', color='E5E7EB')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

print("Criando planilha de Ministerios...")

# ABA 1: Dashboard
ws_inicio = wb.create_sheet("INICIO", 0)

ws_inicio.merge_cells('A1:H1')
ws_inicio['A1'] = 'GESTAO DE MINISTERIOS'
ws_inicio['A1'].font = Font(name='Calibri', size=22, bold=True, color=AZUL_ESCURO)
ws_inicio['A1'].alignment = center_alignment

ws_inicio.merge_cells('A2:H2')
ws_inicio['A2'] = f'Relatorio gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}'
ws_inicio['A2'].font = Font(name='Calibri', size=10, color=CINZA, italic=True)
ws_inicio['A2'].alignment = center_alignment

# Cards de resumo
blue_fill = PatternFill(start_color='EFF6FF', end_color='EFF6FF', fill_type='solid')
green_fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')
gold_fill = PatternFill(start_color='FFFBEB', end_color='FFFBEB', fill_type='solid')

# Card 1
for row in range(4, 7):
    for col in range(1, 4):
        ws_inicio.cell(row=row, column=col).fill = blue_fill
        ws_inicio.cell(row=row, column=col).border = thin_border
ws_inicio.merge_cells('A4:C4')
ws_inicio['A4'] = 'Total de Ministerios'
ws_inicio['A4'].font = Font(size=12, bold=True, color=AZUL_ESCURO)
ws_inicio['A4'].alignment = center_alignment
ws_inicio.merge_cells('A6:C6')
ws_inicio['A6'] = 12
ws_inicio['A6'].font = Font(size=28, bold=True, color=VERDE)
ws_inicio['A6'].alignment = center_alignment

# Card 2
for row in range(4, 7):
    for col in range(4, 7):
        ws_inicio.cell(row=row, column=col).fill = green_fill
        ws_inicio.cell(row=row, column=col).border = thin_border
ws_inicio.merge_cells('D4:F4')
ws_inicio['D4'] = 'Membros Ativos'
ws_inicio['D4'].font = Font(size=12, bold=True, color=AZUL_ESCURO)
ws_inicio['D4'].alignment = center_alignment
ws_inicio.merge_cells('D6:F6')
ws_inicio['D6'] = 387
ws_inicio['D6'].font = Font(size=28, bold=True, color=AZUL_ESCURO)
ws_inicio['D6'].alignment = center_alignment

# Card 3
for row in range(4, 7):
    for col in range(7, 10):
        ws_inicio.cell(row=row, column=col).fill = gold_fill
        ws_inicio.cell(row=row, column=col).border = thin_border
ws_inicio.merge_cells('G4:I4')
ws_inicio['G4'] = 'Orcamento Total'
ws_inicio['G4'].font = Font(size=12, bold=True, color=AZUL_ESCURO)
ws_inicio['G4'].alignment = center_alignment
ws_inicio.merge_cells('G6:I6')
ws_inicio['G6'] = 'R$ 26.300'
ws_inicio['G6'].font = Font(size=28, bold=True, color=DOURADO)
ws_inicio['G6'].alignment = center_alignment

# Tabela de ministerios
ws_inicio['A9'] = 'VISAo GERAL DOS MINISTERIOS'
ws_inicio['A9'].font = Font(name='Calibri', size=14, bold=True, color=AZUL_ESCURO)
ws_inicio.merge_cells('A9:I9')

headers = ['Icone', 'Ministerio', 'Lider', 'Membros', 'Reunioes', 'Status', 'Orcamento']
for col, header in enumerate(headers, 1):
    cell = ws_inicio.cell(row=10, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

ministerios = [
    ('🎨', 'Artes', 'Maria Silva', 15, 4, 'Ativo', 1800),
    ('💑', 'Casais', 'Pastor Roberto', 38, 6, 'Ativo', 1500),
    ('📚', 'Ensino', 'Professora Ana', 25, 8, 'Ativo', 2200),
    ('⚽', 'Esportes', 'Carlos Mendes', 20, 3, 'Ativo', 2000),
    ('🎵', 'Louvor', 'Pastor Joao', 45, 12, 'Ativo', 3500),
    ('🙏', 'Intercessao', 'Irma Ana Costa', 32, 24, 'Ativo', 1200),
    ('📢', 'Evangelismo', 'Pr. Carlos', 28, 6, 'Ativo', 2800),
    ('👶', 'Infantil', 'Tia Maria Joao', 25, 8, 'Ativo', 2200),
    ('🎯', 'Jovens', 'Pr. Ricardo', 52, 10, 'Ativo', 3000),
    ('🏥', 'Diaconia', 'Diacono Jose', 15, 4, 'Ativo', 5000),
    ('📱', 'Midia', 'Joao Vitor', 12, 6, 'Ativo', 4500),
    ('💃', 'Danca', 'Prof. Silvia', 20, 5, 'Ativo', 1800),
]

for row_idx, (icone, nome, lider, membros, reunioes, status, orcamento) in enumerate(ministerios, 11):
    ws_inicio.cell(row=row_idx, column=1, value=icone).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=2, value=nome).font = Font(bold=True, color=AZUL_ESCURO)
    ws_inicio.cell(row=row_idx, column=3, value=lider)
    ws_inicio.cell(row=row_idx, column=4, value=membros).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=5, value=reunioes).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=6, value=status).alignment = center_alignment
    ws_inicio.cell(row=row_idx, column=7, value=f'R$ {orcamento:,}').alignment = center_alignment
    
    for col in range(1, 8):
        ws_inicio.cell(row=row_idx, column=col).border = thin_border

# Ajustar larguras
ws_inicio.column_dimensions['A'].width = 8
ws_inicio.column_dimensions['B'].width = 18
ws_inicio.column_dimensions['C'].width = 22
ws_inicio.column_dimensions['D'].width = 10
ws_inicio.column_dimensions['E'].width = 12
ws_inicio.column_dimensions['F'].width = 12
ws_inicio.column_dimensions['G'].width = 15

# ABA 2: Cadastro Completo
ws_cad = wb.create_sheet("Cadastro")
ws_cad.merge_cells('A1:K1')
ws_cad['A1'] = 'CADASTRO COMPLETO DE MINISTERIOS'
ws_cad['A1'].font = title_font
ws_cad['A1'].alignment = center_alignment

headers_cad = ['Codigo', 'Icone', 'Nome', 'Lider', 'Vice-Lider', 'Membros', 'Fundacao', 'Status', 'Orcamento Mensal', 'Descricao', 'Objetivos']
for col, header in enumerate(headers_cad, 1):
    cell = ws_cad.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

descricoes = [
    'Ministerio de artes visuais, teatro e expressao criativa',
    'Encontros e atividades para casais da igreja',
    'Escola Biblica Dominical e formacao teologica',
    'Atividades esportivas e recreativas para a comunidade',
    'Louvor, adoracao e musica nos cultos',
    'Oracao intercessoria 24 horas pela igreja',
    'Evangelizacao, missoes e novos convertidos',
    'Ministerio infantil e criancas',
    'Juventude e adolescentes',
    'Assistencia social e ajuda aos necessitados',
    'Transmissoes, fotos, videos e redes sociais',
    'Danca, coreografias e apresentacoes',
]

objetivos = [
    'Desenvolver talentos artisticos para Deus',
    'Fortalecer o casamento e familia',
    'Ensinar a Palavra de Deus de forma didatica',
    'Promover saude e comunhao atraves do esporte',
    'Levar a igreja a presenca de Deus',
    'Interceder pela igreja, cidade e nacoes',
    'Ganhar almas para o Reino de Deus',
    'Ensinar as criancas nos caminhos do Senhor',
    'Discipular jovens para Cristo',
    'Praticar o amor ao proximo',
    'Comunicar a mensagem da igreja',
    'Adorar a Deus atraves da danca',
]

for row_idx, ((icone, nome, lider, membros, reunioes, status, orcamento), desc, obj) in enumerate(zip(ministerios, descricoes, objetivos), 4):
    ws_cad.cell(row=row_idx, column=1, value=f'M{row_idx-3:03d}').alignment = center_alignment
    ws_cad.cell(row=row_idx, column=2, value=icone).alignment = center_alignment
    ws_cad.cell(row=row_idx, column=3, value=nome).font = Font(bold=True)
    ws_cad.cell(row=row_idx, column=4, value=lider)
    ws_cad.cell(row=row_idx, column=5, value=f'Vice {nome}')
    ws_cad.cell(row=row_idx, column=6, value=membros).alignment = center_alignment
    ws_cad.cell(row=row_idx, column=7, value=f'201{random.randint(0, 9)}-{random.randint(1, 12):02d}-15').alignment = center_alignment
    ws_cad.cell(row=row_idx, column=8, value=status).alignment = center_alignment
    ws_cad.cell(row=row_idx, column=9, value=orcamento).number_format = 'R$ #,##0'
    ws_cad.cell(row=row_idx, column=9).alignment = center_alignment
    ws_cad.cell(row=row_idx, column=10, value=desc).font = Font(size=9)
    ws_cad.cell(row=row_idx, column=11, value=obj).font = Font(size=9)
    
    for col in range(1, 12):
        cell = ws_cad.cell(row=row_idx, column=col)
        cell.border = thin_border
        cell.fill = PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid')

widths_cad = [10, 8, 15, 22, 20, 10, 12, 10, 18, 40, 35]
for col, width in enumerate(widths_cad, 1):
    ws_cad.column_dimensions[chr(64 + col)].width = width

# ABA 3: Financeiro
ws_fin = wb.create_sheet("Financeiro")
ws_fin.merge_cells('A1:H1')
ws_fin['A1'] = 'CONTROLE FINANCEIRO POR MINISTERIO'
ws_fin['A1'].font = title_font
ws_fin['A1'].alignment = center_alignment

headers_fin = ['Ministerio', 'Orcamento Mensal', 'Gasto Mes', 'Saldo', '% Executado', 'Status Orcamentario']
for col, header in enumerate(headers_fin, 1):
    cell = ws_fin.cell(row=3, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_alignment
    cell.border = thin_border

for row_idx, (icone, nome, lider, membros, reunioes, status, orcamento) in enumerate(ministerios, 4):
    gasto = orcamento * random.uniform(0.6, 0.95)
    saldo = orcamento - gasto
    pct = gasto / orcamento
    
    ws_fin.cell(row=row_idx, column=1, value=f'{icone} {nome}')
    ws_fin.cell(row=row_idx, column=2, value=orcamento).number_format = 'R$ #,##0'
    ws_fin.cell(row=row_idx, column=2).alignment = center_alignment
    ws_fin.cell(row=row_idx, column=3, value=gasto).number_format = 'R$ #,##0'
    ws_fin.cell(row=row_idx, column=3).alignment = center_alignment
    ws_fin.cell(row=row_idx, column=4, value=saldo).number_format = 'R$ #,##0'
    ws_fin.cell(row=row_idx, column=4).alignment = center_alignment
    ws_fin.cell(row=row_idx, column=5, value=pct).number_format = '0%'
    ws_fin.cell(row=row_idx, column=5).alignment = center_alignment
    
    status_orc = 'OK' if pct < 0.9 else 'Atencao' if pct < 1 else 'Estourado'
    ws_fin.cell(row=row_idx, column=6, value=status_orc).alignment = center_alignment
    
    for col in range(1, 7):
        ws_fin.cell(row=row_idx, column=col).border = thin_border

widths_fin = [25, 18, 15, 15, 14, 22]
for col, width in enumerate(widths_fin, 1):
    ws_fin.column_dimensions[chr(64 + col)].width = width

# Salvar na pasta public
nome_arquivo = 'planilha-ministerios.xlsx'
caminho = f'C:/Users/eduka/Desktop/Gestão Igreja/public/{nome_arquivo}'

wb.save(caminho)

print(f"\n{'='*60}")
print(f"[SUCESSO] PLANILHA CRIADA NA PASTA PUBLIC!")
print(f"{'='*60}")
print(f"Arquivo: {nome_arquivo}")
print(f"Local: {caminho}")
print(f"\nABAS CRIADAS:")
print(f"  • INICIO - Dashboard com cards")
print(f"  • Cadastro - Dados completos")
print(f"  • Financeiro - Controle orcamentario")
print(f"\n{'='*60}")
